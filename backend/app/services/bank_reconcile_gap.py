"""Bank-scoped net and Gap helpers (read-only reconcile math for Step 3+)."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

import openpyxl
from sqlalchemy import and_, func, select
from sqlalchemy.orm import Session

from app.models.deposit import Deposit
from app.models.expense import Expense
from app.services.transaction_filters import (
    deposit_company_float_clause,
    expense_company_float_clauses,
)


def _deposit_bank_scoped_clauses(*, after: date | None, verified_only: bool):
    clauses = [
        deposit_company_float_clause(),
        Deposit.bank_reconcile_exclude.is_(False),
        Deposit.transaction_date.is_not(None),
    ]
    if after is not None:
        clauses.append(Deposit.transaction_date > after)
    if verified_only:
        clauses.append(Deposit.bank_verified_at.is_not(None))
    return clauses


def _expense_bank_scoped_clauses(*, after: date | None, verified_only: bool):
    clauses = [
        *expense_company_float_clauses(),
        Expense.bank_reconcile_exclude.is_(False),
        Expense.transaction_date.is_not(None),
        # Paid-by-card merchants verify via CC Excel (Step 6); settlement is Step 7
        Expense.payment_method != "credit_card",
    ]
    if after is not None:
        clauses.append(Expense.transaction_date > after)
    if verified_only:
        clauses.append(Expense.bank_verified_at.is_not(None))
    return clauses


def sum_bank_scoped_nets(
    db: Session,
    *,
    after_date: date | None,
    date_to: date | None = None,
) -> tuple[Decimal, Decimal, Decimal, Decimal]:
    """Return (all_net, verified_net, all_deposits, all_expenses).

    Net = deposits − expenses for bank-scoped rows with transaction_date > after_date
    (and ≤ date_to when provided).
    """

    def _sums(verified_only: bool) -> tuple[Decimal, Decimal]:
        dep_clauses = _deposit_bank_scoped_clauses(
            after=after_date, verified_only=verified_only
        )
        exp_clauses = _expense_bank_scoped_clauses(
            after=after_date, verified_only=verified_only
        )
        if date_to is not None:
            dep_clauses.append(Deposit.transaction_date <= date_to)
            exp_clauses.append(Expense.transaction_date <= date_to)

        dep_total = db.scalar(
            select(func.coalesce(func.sum(Deposit.amount), 0)).where(and_(*dep_clauses))
        )
        exp_total = db.scalar(
            select(func.coalesce(func.sum(Expense.amount), 0)).where(and_(*exp_clauses))
        )
        return Decimal(str(dep_total or 0)), Decimal(str(exp_total or 0))

    all_dep, all_exp = _sums(False)
    ver_dep, ver_exp = _sums(True)
    return all_dep - all_exp, ver_dep - ver_exp, all_dep, all_exp


def _parse_amount(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    # datetime is a subclass of date — normalize so APIs return YYYY-MM-DD only
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d.%m.%Y", "%Y-%m-%d", "%d/%m/%y", "%d.%m.%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_bank_statement_balance(content: bytes) -> dict[str, Any]:
    """Extract B = latest row היתרה בש״ח from a Leumi-style bank Excel.

    Also returns statement_end_date (latest תאריך) and movement row count.
    """
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        header_row = None
        headers: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            vals = [str(v).strip() if v is not None else "" for v in row]
            if "תאריך" in vals and ("בחובה" in vals or "בזכות" in vals):
                header_row = i
                headers = vals
                break
        if header_row is None:
            raise ValueError("Could not find bank movement header row (תאריך / בחובה / בזכות)")

        # Normalize header keys (quote variants on היתרה בש״ח)
        col: dict[str, int] = {}
        balance_idx = None
        for idx, name in enumerate(headers):
            if not name:
                continue
            col[name] = idx
            compact = name.replace('"', "").replace("״", "").replace("'", "")
            if "היתרה" in compact and "ש" in compact:
                balance_idx = idx

        if balance_idx is None:
            for key in ("היתרה בש\"ח", 'היתרה בש"ח', "היתרה בש״ח", "היתרה"):
                if key in col:
                    balance_idx = col[key]
                    break

        date_idx = col.get("תאריך")
        last_balance: Decimal | None = None
        first_date: date | None = None
        last_date: date | None = None
        rows = 0

        for row_number, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_number <= header_row:
                continue
            values = list(row)
            if date_idx is not None and date_idx < len(values):
                tx_date = _parse_date(values[date_idx])
            else:
                tx_date = None

            bal = None
            if balance_idx is not None and balance_idx < len(values):
                bal = _parse_amount(values[balance_idx])

            # Skip empty trailing rows
            if tx_date is None and bal is None:
                nonempty = any(v is not None and str(v).strip() not in ("", "0") for v in values)
                if not nonempty:
                    continue

            rows += 1
            if bal is not None:
                last_balance = bal
            if tx_date is not None:
                if first_date is None or tx_date < first_date:
                    first_date = tx_date
                if last_date is None or tx_date > last_date:
                    last_date = tx_date

        if last_balance is None:
            raise ValueError("No היתרה בש״ח values found in bank file rows")

        return {
            "bank_balance": last_balance,
            "statement_start_date": first_date,
            "statement_end_date": last_date,
            "movement_row_count": rows,
        }
    finally:
        wb.close()


def parse_bank_statement_lines(content: bytes) -> dict[str, Any]:
    """Parse movement rows for reconcile matching (plus balance summary)."""
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        header_row = None
        headers: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            vals = [str(v).strip() if v is not None else "" for v in row]
            if "תאריך" in vals and ("בחובה" in vals or "בזכות" in vals):
                header_row = i
                headers = vals
                break
        if header_row is None:
            raise ValueError("Could not find bank movement header row (תאריך / בחובה / בזכות)")

        col: dict[str, int] = {}
        balance_idx = None
        for idx, name in enumerate(headers):
            if not name:
                continue
            col[name] = idx
            compact = name.replace('"', "").replace("״", "").replace("'", "")
            if "היתרה" in compact and "ש" in compact:
                balance_idx = idx
        if balance_idx is None:
            for key in ("היתרה בש\"ח", 'היתרה בש"ח', "היתרה בש״ח", "היתרה"):
                if key in col:
                    balance_idx = col[key]
                    break

        def get(row_vals: list[Any], name: str) -> Any:
            idx = col.get(name)
            if idx is None or idx >= len(row_vals):
                return None
            return row_vals[idx]

        lines: list[dict[str, Any]] = []
        last_balance: Decimal | None = None
        first_date: date | None = None
        last_date: date | None = None

        for row_number, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_number <= header_row:
                continue
            values = list(row)
            tx_date = _parse_date(get(values, "תאריך"))
            debit = _parse_amount(get(values, "בחובה"))
            credit = _parse_amount(get(values, "בזכות"))
            raw_debit = get(values, "בחובה")
            raw_credit = get(values, "בזכות")
            try:
                if raw_debit is not None and Decimal(str(raw_debit)) == 0:
                    debit = None
            except (InvalidOperation, ValueError):
                pass
            try:
                if raw_credit is not None and Decimal(str(raw_credit)) == 0:
                    credit = None
            except (InvalidOperation, ValueError):
                pass

            asmachta = None
            raw_ref = get(values, "אסמכתא")
            if raw_ref is not None and str(raw_ref).strip():
                asmachta = str(raw_ref).strip()
            desc = get(values, "תיאור")
            extended = get(values, "תאור מורחב")
            desc_s = str(desc).strip() if desc is not None else ""
            ext_s = str(extended).strip() if extended is not None else ""
            full_desc = " | ".join(p for p in (desc_s, ext_s) if p) or None

            bal = None
            if balance_idx is not None and balance_idx < len(values):
                bal = _parse_amount(values[balance_idx])

            if tx_date is None and debit is None and credit is None and not full_desc and not asmachta:
                continue
            if debit is None and credit is None:
                continue

            if credit is not None:
                side = "credit"
                amount = credit
            else:
                side = "debit"
                amount = debit
            assert amount is not None

            fingerprint = "|".join(
                [
                    str(row_number),
                    tx_date.isoformat() if tx_date else "",
                    side,
                    f"{amount:.2f}",
                    asmachta or "",
                ]
            )
            lines.append(
                {
                    "fingerprint": fingerprint,
                    "row_number": row_number,
                    "transaction_date": tx_date.isoformat() if tx_date else None,
                    "side": side,
                    "amount": str(amount),
                    "asmachta": asmachta,
                    "description": full_desc,
                    "balance_after": str(bal) if bal is not None else None,
                    "status": "unmatched",
                    "proposed_kind": None,
                    "proposed_tx_id": None,
                    "proposed_tx_ref": None,
                    "proposed_summary": None,
                    "match_confidence": None,
                    "ignore_reason": None,
                }
            )
            if bal is not None:
                last_balance = bal
            if tx_date is not None:
                if first_date is None or tx_date < first_date:
                    first_date = tx_date
                if last_date is None or tx_date > last_date:
                    last_date = tx_date

        if last_balance is None:
            raise ValueError("No היתרה בש״ח values found in bank file rows")
        if not lines:
            raise ValueError("No bank movement amounts found in file")

        return {
            "bank_balance": last_balance,
            "statement_start_date": first_date,
            "statement_end_date": last_date,
            "movement_row_count": len(lines),
            "lines": lines,
        }
    finally:
        wb.close()
