"""Credit-card Excel match / verify (Step 6) — match paid-by-card expenses, no mass create."""

from __future__ import annotations

import copy
import re
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
from uuid import UUID

import openpyxl
from sqlalchemy import and_, select
from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified

from app.models.cc_reconcile_session import CcReconcileSession
from app.models.expense import Expense
from app.models.property import Property


def _parse_amount(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d.%m.%Y", "%Y-%m-%d", "%d/%m/%y", "%d.%m.%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _optional_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_cc_statement_lines(content: bytes) -> dict[str, Any]:
    """Parse Leumi-style credit-card Excel into charge lines for matching."""
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        card_last4 = "unknown"
        for row in ws.iter_rows(values_only=True, max_row=10):
            for cell in row:
                if cell is None:
                    continue
                text = str(cell)
                match = re.search(r"(\d{4})\s*$", text)
                if match and ("מסטרקארד" in text or "כרטיס" in text or "מאסטרקארד" in text):
                    card_last4 = match.group(1)
                    break

        header_row = None
        headers: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            vals = [str(v).strip() if v is not None else "" for v in row]
            if "תאריך העסקה" in vals and "סכום חיוב" in vals:
                header_row = i
                headers = vals
                break
        if header_row is None:
            raise ValueError("Could not find credit-card header row (תאריך העסקה / סכום חיוב)")

        col = {name: idx for idx, name in enumerate(headers) if name}

        def get(row_vals: list[Any], name: str) -> Any:
            idx = col.get(name)
            if idx is None or idx >= len(row_vals):
                return None
            return row_vals[idx]

        lines: list[dict[str, Any]] = []
        first_date: date | None = None
        last_date: date | None = None

        for row_number, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_number <= header_row:
                continue
            values = list(row)
            if any(v is not None and "סה" in str(v) for v in values[:5]):
                continue

            tx_date = _parse_date(get(values, "תאריך העסקה"))
            merchant = _optional_str(get(values, "שם בית העסק"))
            charge = _parse_amount(get(values, "סכום חיוב"))
            details = _optional_str(get(values, "פרטים"))
            if charge is None and tx_date is None and merchant is None:
                continue
            if charge is None:
                continue
            # Positive charge = expense on card; skip credits (negative) for Stage A match
            if charge <= 0:
                continue

            amount = charge
            fingerprint = "|".join(
                [
                    str(row_number),
                    tx_date.isoformat() if tx_date else "",
                    f"{amount:.2f}",
                    merchant or "",
                ]
            )
            lines.append(
                {
                    "fingerprint": fingerprint,
                    "row_number": row_number,
                    "transaction_date": tx_date.isoformat() if tx_date else None,
                    "amount": str(amount),
                    "merchant": merchant,
                    "details": details,
                    "status": "unmatched",
                    "proposed_tx_id": None,
                    "proposed_tx_ref": None,
                    "proposed_summary": None,
                    "match_confidence": None,
                    "ignore_reason": None,
                }
            )
            if tx_date is not None:
                if first_date is None or tx_date < first_date:
                    first_date = tx_date
                if last_date is None or tx_date > last_date:
                    last_date = tx_date

        if not lines:
            raise ValueError("No credit-card charge rows found in file")

        return {
            "card_last4": card_last4,
            "statement_start_date": first_date,
            "statement_end_date": last_date,
            "movement_row_count": len(lines),
            "lines": lines,
        }
    finally:
        wb.close()


def _cc_pending_filters(*, date_from: date | None, date_to: date | None):
    clauses = [
        Expense.payment_method == "credit_card",
        Expense.cc_verified_at.is_(None),
        Expense.bank_reconcile_exclude.is_(False),
        Expense.amount > 0,
        Expense.transaction_date.is_not(None),
    ]
    if date_from is not None:
        clauses.append(Expense.transaction_date >= date_from)
    if date_to is not None:
        clauses.append(Expense.transaction_date <= date_to)
    return clauses


def _score_cc(line: dict, row: Expense) -> int:
    amount = Decimal(str(line["amount"]))
    if row.amount != amount:
        return -1
    score = 2
    line_date = date.fromisoformat(line["transaction_date"]) if line.get("transaction_date") else None
    if line_date and row.transaction_date == line_date:
        score += 3
    elif line_date and row.transaction_date and abs((row.transaction_date - line_date).days) <= 2:
        score += 1
    merchant = (line.get("merchant") or "").lower()
    hay = " ".join(
        p for p in (row.vendor_name or "", row.description or "", row.category or "") if p
    ).lower()
    if merchant and hay:
        if merchant in hay or hay in merchant:
            score += 4
        else:
            # token overlap
            m_tokens = {t for t in re.split(r"\W+", merchant) if len(t) > 2}
            h_tokens = {t for t in re.split(r"\W+", hay) if len(t) > 2}
            if m_tokens & h_tokens:
                score += 2
    return score


def _propose_matches(db: Session, lines: list[dict], *, date_from: date | None, date_to: date | None) -> None:
    candidates = list(
        db.scalars(select(Expense).where(and_(*_cc_pending_filters(date_from=date_from, date_to=date_to))))
    )
    used: set[str] = set()
    for line in lines:
        best: Expense | None = None
        best_score = 0
        for row in candidates:
            rid = str(row.id)
            if rid in used:
                continue
            score = _score_cc(line, row)
            if score > best_score:
                best_score = score
                best = row
        # Require amount match + at least date or merchant signal
        if best is not None and best_score >= 5:
            used.add(str(best.id))
            line["status"] = "proposed_match"
            line["proposed_tx_id"] = str(best.id)
            line["proposed_tx_ref"] = best.transaction_ref
            line["proposed_summary"] = best.vendor_name or best.description or best.category
            line["match_confidence"] = "high" if best_score >= 8 else "medium"


def _unmatched_cc_app(
    db: Session,
    *,
    date_from: date | None,
    date_to: date | None,
    matched_ids: set[str],
) -> list[dict]:
    rows = list(
        db.scalars(select(Expense).where(and_(*_cc_pending_filters(date_from=date_from, date_to=date_to))))
    )
    out: list[dict] = []
    for row in rows:
        rid = str(row.id)
        if rid in matched_ids:
            continue
        out.append(
            {
                "kind": "expense",
                "id": rid,
                "transaction_ref": row.transaction_ref,
                "transaction_date": row.transaction_date.isoformat() if row.transaction_date else None,
                "amount": str(row.amount),
                "description": row.vendor_name or row.description or row.category,
                "status": "unmatched",
                "ignore_reason": None,
            }
        )
    out.sort(key=lambda r: (r.get("transaction_date") or "", r["id"]))
    return out


def create_session_from_upload(
    db: Session, *, content: bytes, filename: str | None
) -> CcReconcileSession:
    parsed = parse_cc_statement_lines(content)
    lines = parsed["lines"]
    date_from = parsed["statement_start_date"]
    date_to = parsed["statement_end_date"]
    _propose_matches(db, lines, date_from=date_from, date_to=date_to)
    matched_ids = {
        line["proposed_tx_id"]
        for line in lines
        if line.get("status") == "proposed_match" and line.get("proposed_tx_id")
    }
    unmatched_app = _unmatched_cc_app(
        db, date_from=date_from, date_to=date_to, matched_ids=matched_ids
    )
    session = CcReconcileSession(
        status="in_progress",
        filename=filename,
        card_last4=parsed["card_last4"],
        statement_start_date=date_from,
        statement_end_date=date_to,
        lines_json=lines,
        unmatched_app_json=unmatched_app,
    )
    db.add(session)
    db.commit()
    db.refresh(session)
    return session


def session_summary(_db: Session, session: CcReconcileSession) -> dict:
    lines = list(session.lines_json or [])
    apps = list(session.unmatched_app_json or [])
    counts = {
        "proposed_match": 0,
        "matched": 0,
        "ignored": 0,
        "unmatched": 0,
        "added": 0,
    }
    for line in lines:
        st = line.get("status") or "unmatched"
        counts[st] = counts.get(st, 0) + 1
    app_unmatched = sum(1 for a in apps if a.get("status") == "unmatched")
    app_ignored = sum(1 for a in apps if a.get("status") == "ignored")
    unresolved_cc = sum(
        1 for line in lines if line.get("status") in ("unmatched", "proposed_match")
    )
    unresolved_app = app_unmatched
    can_complete = unresolved_cc == 0 and unresolved_app == 0
    return {
        "id": str(session.id),
        "status": session.status,
        "filename": session.filename,
        "card_last4": session.card_last4,
        "statement_start_date": session.statement_start_date.isoformat()
        if session.statement_start_date
        else None,
        "statement_end_date": session.statement_end_date.isoformat()
        if session.statement_end_date
        else None,
        "counts": {
            **counts,
            "app_unmatched": app_unmatched,
            "app_ignored": app_ignored,
            "unresolved_cc": unresolved_cc,
            "unresolved_app": unresolved_app,
        },
        "can_complete": can_complete,
        "lines": lines,
        "unmatched_app": apps,
    }


def apply_actions(db: Session, session: CcReconcileSession, actions: list[dict]) -> CcReconcileSession:
    if session.status != "in_progress":
        raise ValueError("Session is not in progress")
    lines = {line["fingerprint"]: line for line in (session.lines_json or [])}
    apps = {f"{a['kind']}:{a['id']}": a for a in (session.unmatched_app_json or [])}
    now = datetime.now(timezone.utc)

    for action in actions:
        kind = action.get("action")
        if kind == "confirm_match":
            fp = action["fingerprint"]
            line = lines.get(fp)
            if not line:
                raise ValueError(f"Unknown CC line {fp}")
            tx_id = action.get("tx_id") or line.get("proposed_tx_id")
            if not tx_id:
                raise ValueError("confirm_match requires tx_id")
            row = db.get(Expense, UUID(str(tx_id)))
            if not row:
                raise ValueError(f"Expense {tx_id} not found")
            if row.payment_method != "credit_card":
                raise ValueError("Only paid-by-card expenses can be CC-verified")
            row.cc_verified_at = now
            line["status"] = "matched"
            line["proposed_tx_id"] = str(tx_id)
            line["proposed_tx_ref"] = row.transaction_ref
            apps.pop(f"expense:{tx_id}", None)

        elif kind == "ignore_cc":
            fp = action["fingerprint"]
            line = lines.get(fp)
            if not line:
                raise ValueError(f"Unknown CC line {fp}")
            reason = (action.get("reason") or "").strip() or "Ignored"
            line["status"] = "ignored"
            line["ignore_reason"] = reason

        elif kind == "ignore_app":
            tx_id = str(action["tx_id"])
            key = f"expense:{tx_id}"
            app = apps.get(key)
            if not app:
                continue
            reason = (action.get("reason") or "").strip() or "Ignored"
            app["status"] = "ignored"
            app["ignore_reason"] = reason

        elif kind == "add_from_cc":
            fp = action["fingerprint"]
            line = lines.get(fp)
            if not line:
                raise ValueError(f"Unknown CC line {fp}")
            property_id = action.get("property_id")
            if not property_id:
                raise ValueError("add_from_cc requires property_id")
            prop = db.get(Property, UUID(str(property_id)))
            if not prop:
                raise ValueError("Property not found")
            amount = Decimal(str(line["amount"]))
            tx_date = (
                date.fromisoformat(line["transaction_date"])
                if line.get("transaction_date")
                else None
            )
            merchant = line.get("merchant")
            row = Expense(
                property_id=prop.id,
                transaction_date=tx_date,
                amount=amount,
                currency="ILS",
                category=(merchant[:255] if merchant else "credit_card"),
                source="credit_card",
                payment_method="credit_card",
                vendor_name=merchant,
                description=line.get("details") or merchant,
                cc_verified_at=now,
            )
            db.add(row)
            db.flush()
            line["status"] = "added"
            line["proposed_tx_id"] = str(row.id)
            line["proposed_tx_ref"] = row.transaction_ref
        else:
            raise ValueError(f"Unknown action {kind}")

    session.lines_json = copy.deepcopy(list(lines.values()))
    session.unmatched_app_json = copy.deepcopy(list(apps.values()))
    flag_modified(session, "lines_json")
    flag_modified(session, "unmatched_app_json")
    db.add(session)
    db.commit()
    db.refresh(session)
    return session


def complete_session(db: Session, session: CcReconcileSession) -> CcReconcileSession:
    summary = session_summary(db, session)
    if not summary["can_complete"]:
        raise ValueError(
            "Cannot complete: unresolved CC file lines or unpaid-by-card app rows remain"
        )
    session.status = "completed"
    db.add(session)
    db.commit()
    db.refresh(session)
    return session
