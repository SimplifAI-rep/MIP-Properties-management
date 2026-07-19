"""Import client Excel workbooks into SimplifAI.

Supported sources (data/ClientData):
- client list to print.xlsx  → owners + properties
- Management expenses sheet.xlsx → per-property expenses + inflows (deposits)
- Bank Account example.xlsx → company bank statement rows
- credit card 1/2 example.xlsx → credit-card expenses

Designed for idempotent re-runs via import_key on deposits/expenses.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import PROJECT_ROOT, get_settings
from app.models.bank_account import BankAccount
from app.models.deposit import Deposit
from app.models.expense import Expense
from app.models.owner import Owner
from app.models.property import Property

logger = logging.getLogger(__name__)

CLIENT_DATA_DIR = PROJECT_ROOT / "data" / "ClientData"
CLIENT_LIST_FILE = "client list to print.xlsx"
MANAGEMENT_FILE = "Management expenses sheet.xlsx"
BANK_FILE = "Bank Account example.xlsx"
CREDIT_CARD_FILES = ("credit card 1 example.xlsx", "credit card 2 example.xlsx")

META_SHEETS = {
    "prop id",
    "dashboard",
    "buffer",
    "oldbuffer",
    "expenses",
    "bills",
}

BUFFER_PROP_ID = "BUFFER"
COMPANY_OWNER_NAME = "My Israel Property (MIP)"
COMPANY_ACCOUNT_NUMBER = "MIP-LEUMI-OPS"
COMPANY_CC_ACCOUNT_PREFIX = "MIP-LEUMI-CC-"


@dataclass
class SkippedRow:
    source_file: str
    sheet: str
    row_number: int
    reason: str
    prop_id: str | None = None
    transaction_date: str | None = None
    amount: str | None = None
    inflow: str | None = None
    section: str | None = None
    notes: str | None = None
    details: str | None = None
    import_key: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_file": self.source_file,
            "sheet": self.sheet,
            "row_number": self.row_number,
            "reason": self.reason,
            "prop_id": self.prop_id,
            "transaction_date": self.transaction_date,
            "amount": self.amount,
            "inflow": self.inflow,
            "section": self.section,
            "notes": self.notes,
            "details": self.details,
            "import_key": self.import_key,
        }


@dataclass
class ImportStats:
    owners_created: int = 0
    properties_created: int = 0
    expenses_created: int = 0
    expenses_skipped: int = 0
    deposits_created: int = 0
    deposits_skipped: int = 0
    bank_accounts_created: int = 0
    rows_seen: int = 0
    rows_skipped_empty: int = 0
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    sheet_counts: dict[str, dict[str, int]] = field(default_factory=dict)
    skipped_rows: list[SkippedRow] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "owners_created": self.owners_created,
            "properties_created": self.properties_created,
            "expenses_created": self.expenses_created,
            "expenses_skipped": self.expenses_skipped,
            "deposits_created": self.deposits_created,
            "deposits_skipped": self.deposits_skipped,
            "bank_accounts_created": self.bank_accounts_created,
            "rows_seen": self.rows_seen,
            "rows_skipped_empty": self.rows_skipped_empty,
            "skipped_row_count": len(self.skipped_rows),
            "warnings": self.warnings,
            "errors": self.errors,
            "sheet_counts": self.sheet_counts,
        }


def normalize_prop_key(raw: Any) -> str:
    """Normalize client property codes for matching.

    Examples:
      "5 or 05 ex" -> "5"
      "05 ex" -> "5"
      "C4or c4 EX" -> "C4"
      "C6 or 06 ex" -> "C6"
      "N160" -> "N160"
      "p14" -> "P14"
    """
    if raw is None:
        return ""
    text = str(raw).strip()
    if not text:
        return ""

    # "C4or c4 EX" (missing space before or)
    text = re.sub(r"(?i)([A-Za-z0-9])or\s+", r"\1 or ", text)
    primary = re.split(r"(?i)\s+or\s+", text, maxsplit=1)[0].strip()
    primary = re.sub(r"(?i)\s*ex\s*$", "", primary).strip()
    primary = re.sub(r"\s+", "", primary).upper()
    primary = primary.replace('"', "").replace("'", "")

    if primary.isdigit():
        return str(int(primary))
    # C06 -> C6 if trailing digits have leading zeros
    m = re.match(r"^([A-Z]+)0*(\d+)$", primary)
    if m:
        return f"{m.group(1)}{int(m.group(2))}"
    return primary


def prop_key_aliases(raw: Any) -> set[str]:
    """All normalized keys that should resolve to the same property."""
    if raw is None:
        return set()
    text = str(raw).strip()
    if not text:
        return set()

    text = re.sub(r"(?i)([A-Za-z0-9])or\s+", r"\1 or ", text)
    parts = re.split(r"(?i)\s+or\s+", text)
    aliases: set[str] = set()
    for part in parts:
        key = normalize_prop_key(part)
        if key:
            aliases.add(key)
        # also keep EX form stripped already by normalize
        bare = re.sub(r"(?i)\s*ex\s*$", "", part.strip())
        bare_key = normalize_prop_key(bare)
        if bare_key:
            aliases.add(bare_key)
    return aliases


def _optional_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        parsed = datetime.fromisoformat(text.replace("Z", ""))
        return parsed.date()
    except ValueError:
        pass
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d.%m.%y", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_amount(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        amount = value
    else:
        text = str(value).strip().replace(",", "").replace("₪", "").replace("NIS", "")
        if not text or text.lower() in {"none", "nan"}:
            return None
        try:
            amount = Decimal(text)
        except (InvalidOperation, ValueError):
            return None
    if amount == 0:
        return None
    return abs(amount).quantize(Decimal("0.01"))


def _map_payment(method: Any) -> tuple[str, str, str | None]:
    """Return (source, payment_method, extra_reference)."""
    if method is None:
        return "management_ledger", "company_account", None
    text = str(method).strip()
    if not text:
        return "management_ledger", "company_account", None

    lower = text.lower()
    if lower in {"cc", "credit", "credit card", "mastercard"}:
        return "credit_card", "credit_card", None
    if lower in {"transfer", "bank transfer", "העברה"}:
        return "management_ledger", "bank_transfer", None
    if lower in {"cash", "מזומן"}:
        return "management_ledger", "cash", None
    if lower in {"deposit", "deposit "}:
        return "manual_owner", "bank_transfer", None
    if "buffer" in lower:
        return "management_ledger", "company_account", "from_buffer"
    # Numeric bank refs used as method in the ledger
    if re.fullmatch(r"\d{4,}", text):
        return "management_ledger", "bank_transfer", text
    return "management_ledger", "company_account", text


def _owner_display_name(last_name: Any, first_name: Any, hebrew: Any) -> str:
    last = _optional_str(last_name) or ""
    first = _optional_str(first_name) or ""
    heb = _optional_str(hebrew)
    english = " ".join(p for p in (first, last) if p).strip()
    if english and heb:
        return f"{english} ({heb})"
    return english or heb or "Unknown owner"


def _is_truthy_reconciled(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    return text in {"y", "yes", "true", "1", "v", "✓"}


class ClientDataImportService:
    def __init__(self, db: Session, data_dir: Path | None = None):
        self.db = db
        self.data_dir = data_dir or CLIENT_DATA_DIR
        self.settings = get_settings()
        self.stats = ImportStats()
        # alias_key -> client_prop_id (canonical)
        self.alias_to_prop: dict[str, str] = {}
        # client_prop_id -> Property
        self.properties_by_id: dict[str, Property] = {}
        self.existing_expense_keys: set[str] = set()
        self.existing_deposit_keys: set[str] = set()

    def import_all(self, *, include_bank: bool = True, include_credit_cards: bool = True) -> ImportStats:
        self._load_existing_keys()
        self._ensure_company_owner_and_buffer()
        self._import_client_list()
        self._import_management_workbook()
        if include_bank:
            self._import_bank_statement()
        if include_credit_cards:
            self._import_credit_cards()
        self.db.commit()
        return self.stats

    def _load_existing_keys(self) -> None:
        self.existing_expense_keys = {
            k for k in self.db.scalars(select(Expense.import_key)).all() if k
        }
        self.existing_deposit_keys = {
            k for k in self.db.scalars(select(Deposit.import_key)).all() if k
        }
        for prop in self.db.scalars(select(Property)).all():
            self.properties_by_id[prop.client_prop_id] = prop
            self.alias_to_prop[normalize_prop_key(prop.client_prop_id)] = prop.client_prop_id

    def _record_skip(
        self,
        *,
        source_file: str,
        sheet: str,
        row_number: int,
        reason: str,
        prop_id: Any = None,
        transaction_date: Any = None,
        amount: Any = None,
        inflow: Any = None,
        section: Any = None,
        notes: Any = None,
        details: str | None = None,
        import_key: str | None = None,
    ) -> None:
        def _fmt(value: Any) -> str | None:
            if value is None:
                return None
            if isinstance(value, Decimal):
                return str(value)
            if isinstance(value, date):
                return value.isoformat()
            text = str(value).strip()
            return text or None

        self.stats.skipped_rows.append(
            SkippedRow(
                source_file=source_file,
                sheet=sheet,
                row_number=row_number,
                reason=reason,
                prop_id=_fmt(prop_id),
                transaction_date=_fmt(transaction_date),
                amount=_fmt(amount),
                inflow=_fmt(inflow),
                section=_fmt(section),
                notes=_fmt(notes),
                details=details,
                import_key=import_key,
            )
        )

    def _ensure_company_owner_and_buffer(self) -> None:
        owner = self.db.scalars(
            select(Owner).where(Owner.name == COMPANY_OWNER_NAME)
        ).first()
        if not owner:
            owner = Owner(name=COMPANY_OWNER_NAME)
            self.db.add(owner)
            self.db.flush()
            self.stats.owners_created += 1

        buffer = self.properties_by_id.get(BUFFER_PROP_ID)
        if not buffer:
            buffer = self.db.scalars(
                select(Property).where(Property.client_prop_id == BUFFER_PROP_ID)
            ).first()
        if not buffer:
            buffer = Property(
                owner_id=owner.id,
                client_prop_id=BUFFER_PROP_ID,
                name="MIP Company Buffer",
                address="Company float / unallocated",
                city=None,
                status="active",
            )
            self.db.add(buffer)
            self.db.flush()
            self.stats.properties_created += 1

        self.properties_by_id[BUFFER_PROP_ID] = buffer
        self.alias_to_prop[BUFFER_PROP_ID] = BUFFER_PROP_ID

        account = self.db.scalars(
            select(BankAccount).where(BankAccount.account_number == COMPANY_ACCOUNT_NUMBER)
        ).first()
        if not account:
            account = BankAccount(
                property_id=None,
                bank_name="Bank Leumi",
                account_number=COMPANY_ACCOUNT_NUMBER,
                currency=self.settings.default_currency,
                label="MIP operating account",
            )
            self.db.add(account)
            self.db.flush()
            self.stats.bank_accounts_created += 1

    def _register_aliases(self, canonical: str, raw: Any) -> None:
        for alias in prop_key_aliases(raw):
            self.alias_to_prop[alias] = canonical
        self.alias_to_prop[normalize_prop_key(canonical)] = canonical

    def _resolve_property(self, raw: Any) -> Property | None:
        if raw is None:
            return None
        text = str(raw).strip()
        if not text:
            return None

        key = normalize_prop_key(text)
        if not key:
            return None
        # Slash/plus codes do not normalize to a single property
        if "/" in text or "+" in text:
            return None

        canonical = self.alias_to_prop.get(key)
        if not canonical:
            for alias in prop_key_aliases(text):
                canonical = self.alias_to_prop.get(alias)
                if canonical:
                    break
        if not canonical:
            return None
        return self.properties_by_id.get(canonical)

    def _import_client_list(self) -> None:
        path = self.data_dir / CLIENT_LIST_FILE
        if not path.exists():
            self.stats.errors.append(f"Missing client list: {path}")
            return

        wb = openpyxl.load_workbook(path, data_only=True)
        current = wb["current clients"] if "current clients" in wb.sheetnames else wb[wb.sheetnames[0]]
        past_name = next((n for n in wb.sheetnames if n.strip().lower().startswith("past")), None)

        owners_by_name: dict[str, Owner] = {
            o.name: o for o in self.db.scalars(select(Owner)).all()
        }

        def upsert_row(row_values: list[Any], *, status: str) -> None:
            if not row_values:
                return
            prop_raw = row_values[0] if len(row_values) > 0 else None
            last_name = row_values[1] if len(row_values) > 1 else None
            first_name = row_values[2] if len(row_values) > 2 else None
            hebrew = row_values[3] if len(row_values) > 3 else None
            address = row_values[4] if len(row_values) > 4 else None
            city = row_values[5] if len(row_values) > 5 else None

            # skip header
            if prop_raw is None:
                return
            header_probe = str(prop_raw).strip().lower()
            if header_probe in {"property number", "property\nnumber", "#", "prop id"}:
                return

            canonical = normalize_prop_key(prop_raw)
            if not canonical:
                self.stats.warnings.append(f"Skipped client row with empty prop id: {row_values[:3]}")
                self._record_skip(
                    source_file=CLIENT_LIST_FILE,
                    sheet=status,
                    row_number=0,
                    reason="empty_prop_id",
                    details=str(row_values[:6]),
                )
                return

            owner_name = _owner_display_name(last_name, first_name, hebrew)
            owner = owners_by_name.get(owner_name)
            if not owner:
                owner = Owner(name=owner_name)
                self.db.add(owner)
                self.db.flush()
                owners_by_name[owner_name] = owner
                self.stats.owners_created += 1

            existing = self.properties_by_id.get(canonical)
            if existing is None:
                existing = self.db.scalars(
                    select(Property).where(Property.client_prop_id == canonical)
                ).first()

            addr = _optional_str(address)
            city_s = _optional_str(city)
            display_name = f"{addr}" if addr else f"Property {canonical}"
            if city_s and addr:
                display_name = f"{addr}, {city_s}"

            if existing is None:
                existing = Property(
                    owner_id=owner.id,
                    client_prop_id=canonical,
                    name=display_name,
                    address=addr,
                    city=city_s,
                    status=status,
                )
                self.db.add(existing)
                self.db.flush()
                self.stats.properties_created += 1
            else:
                # Keep current list as source of truth when overlapping
                if status == "active" or existing.status != "active":
                    existing.owner_id = owner.id
                    existing.name = display_name
                    existing.address = addr
                    existing.city = city_s
                    existing.status = status

            self.properties_by_id[canonical] = existing
            self._register_aliases(canonical, prop_raw)

        for i, row in enumerate(current.iter_rows(values_only=True), 1):
            if i == 1:
                continue
            upsert_row(list(row), status="active")

        if past_name:
            past = wb[past_name]
            for i, row in enumerate(past.iter_rows(values_only=True), 1):
                if i == 1:
                    continue
                vals = list(row)
                prop_raw = vals[0] if vals else None
                canonical = normalize_prop_key(prop_raw)
                if canonical and canonical in self.properties_by_id:
                    # already active from current list — do not downgrade
                    self._register_aliases(canonical, prop_raw)
                    continue
                upsert_row(vals, status="inactive")

        self.db.flush()

    def _import_management_workbook(self) -> None:
        path = self.data_dir / MANAGEMENT_FILE
        if not path.exists():
            self.stats.errors.append(f"Missing management workbook: {path}")
            return

        wb = openpyxl.load_workbook(path, data_only=True)

        # Import Buffer sheet onto BUFFER property
        if "Buffer" in wb.sheetnames:
            self._import_ledger_sheet(wb["Buffer"], default_prop_id=BUFFER_PROP_ID, sheet_label="Buffer")

        for sheet_name in wb.sheetnames:
            if sheet_name.strip().lower() in META_SHEETS:
                continue
            # Combined multi-property sheet — import with per-row Prop ID
            if sheet_name.strip() in {"801-618- 619", "801-618-619"}:
                self._import_ledger_sheet(wb[sheet_name], default_prop_id=None, sheet_label=sheet_name)
                continue

            prop = self._resolve_property(sheet_name)
            if prop is None:
                # Auto-create property from sheet name so we don't lose history
                canonical = normalize_prop_key(sheet_name)
                if not canonical:
                    self.stats.warnings.append(f"Skipping sheet with unusable name: {sheet_name!r}")
                    continue
                company = self.db.scalars(
                    select(Owner).where(Owner.name == COMPANY_OWNER_NAME)
                ).first()
                assert company is not None
                prop = Property(
                    owner_id=company.id,
                    client_prop_id=canonical,
                    name=f"Property {canonical} (from ledger)",
                    status="active",
                )
                self.db.add(prop)
                self.db.flush()
                self.properties_by_id[canonical] = prop
                self._register_aliases(canonical, sheet_name)
                self.stats.properties_created += 1
                self.stats.warnings.append(
                    f"Created property {canonical} from sheet {sheet_name!r} (not in client list)"
                )

            self._import_ledger_sheet(
                wb[sheet_name],
                default_prop_id=prop.client_prop_id,
                sheet_label=sheet_name,
            )

        self.db.flush()

    def _import_ledger_sheet(
        self,
        ws,
        *,
        default_prop_id: str | None,
        sheet_label: str,
    ) -> None:
        counts = {
            "expenses": 0,
            "deposits": 0,
            "skipped": 0,
            "seen": 0,
            "resident_paid": 0,
            "owner_paid": 0,
            "paid_by_company": 0,
            "nearly_cc": 0,
            "cash": 0,
            "other": 0,
            "rental_income": 0,
        }
        header_idx: dict[str, int] | None = None

        for row_number, row in enumerate(ws.iter_rows(values_only=True), 1):
            values = list(row)
            if not any(v is not None and str(v).strip() != "" for v in values):
                continue

            if header_idx is None:
                mapped = self._detect_ledger_header(values)
                if mapped:
                    header_idx = mapped
                    continue
                # Some sheets may have title rows — keep scanning
                continue

            counts["seen"] += 1
            self.stats.rows_seen += 1

            def col(name: str) -> Any:
                idx = header_idx.get(name)  # type: ignore[union-attr]
                if idx is None or idx >= len(values):
                    return None
                return values[idx]

            prop_raw = col("prop_id")
            prop = self._resolve_property(prop_raw) if prop_raw is not None else None
            if prop is None and default_prop_id:
                prop = self.properties_by_id.get(default_prop_id)
            # Combined codes on sheets without a default property → company buffer
            if prop is None and prop_raw is not None:
                raw_text = str(prop_raw)
                if "/" in raw_text or "+" in raw_text:
                    prop = self.properties_by_id.get(BUFFER_PROP_ID)
            if prop is None:
                counts["skipped"] += 1
                self.stats.rows_skipped_empty += 1
                if len(self.stats.warnings) < 50:
                    self.stats.warnings.append(
                        f"{sheet_label} R{row_number}: could not resolve property {prop_raw!r}"
                    )
                self._record_skip(
                    source_file=MANAGEMENT_FILE,
                    sheet=sheet_label,
                    row_number=row_number,
                    reason="unresolved_property",
                    prop_id=prop_raw,
                    transaction_date=col("date"),
                    amount=col("amount"),
                    inflow=col("inflow"),
                    section=col("section") or col("type"),
                    notes=col("notes"),
                    details="Could not match Prop ID to a property",
                )
                continue

            tx_date = _parse_date(col("date"))
            amount = _parse_amount(col("amount"))
            inflow = _parse_amount(col("inflow"))
            resident_paid_amount = _parse_amount(col("he_she_paid"))
            owner_paid_amount = _parse_amount(col("owner_paid"))
            mip_paid_amount = _parse_amount(col("mip"))
            nearly_cc_amount = _parse_amount(col("nearly_cc"))
            cash_amount = _parse_amount(col("cash"))
            other_amount = _parse_amount(col("other"))
            rental_income_amount = _parse_amount(col("rental_income"))
            section = _optional_str(col("section")) or _optional_str(col("type")) or "other"
            notes = _optional_str(col("notes"))
            vendor = _optional_str(col("company"))
            receipt = _optional_str(col("receipt"))
            method = col("method")
            reconciled = _is_truthy_reconciled(col("reconciled"))

            if tx_date is None:
                # Starting balance / header-like rows
                reason = (
                    "starting_balance"
                    if section and "starting" in section.lower()
                    else "missing_date"
                )
                counts["skipped"] += 1
                if reason == "missing_date":
                    self.stats.rows_skipped_empty += 1
                self._record_skip(
                    source_file=MANAGEMENT_FILE,
                    sheet=sheet_label,
                    row_number=row_number,
                    reason=reason,
                    prop_id=prop.client_prop_id,
                    transaction_date=col("date"),
                    amount=amount,
                    inflow=inflow,
                    section=section,
                    notes=notes,
                    details="Row has no usable transaction date",
                )
                continue

            # Any money column on the ledger row
            if (
                amount is None
                and inflow is None
                and resident_paid_amount is None
                and owner_paid_amount is None
                and mip_paid_amount is None
                and nearly_cc_amount is None
                and cash_amount is None
                and other_amount is None
                and rental_income_amount is None
            ):
                counts["skipped"] += 1
                self.stats.rows_skipped_empty += 1
                self._record_skip(
                    source_file=MANAGEMENT_FILE,
                    sheet=sheet_label,
                    row_number=row_number,
                    reason="no_money_columns",
                    prop_id=prop.client_prop_id,
                    transaction_date=tx_date,
                    section=section,
                    notes=notes,
                    details="Date present but Amount/Inflow/paid columns are empty",
                )
                continue

            source, payment_method, method_ref = _map_payment(method)
            sheet_slug = re.sub(r"[^A-Za-z0-9]+", "", sheet_label).lower() or "sheet"
            desc_parts = [p for p in (section, notes) if p]
            description = " | ".join(desc_parts) if desc_parts else section

            def add_expense(
                *,
                key_kind: str,
                exp_amount: Decimal,
                exp_source: str,
                exp_method: str,
                paid_by_resident: bool = False,
                paid_by_company: bool = False,
                paid_by_owner: bool = False,
                ledger_column: str | None = None,
                count_key: str | None = None,
            ) -> None:
                import_key = (
                    f"mgmt:{sheet_slug}:r{row_number}:{key_kind}:"
                    f"{tx_date.isoformat()}:{exp_amount}"
                )
                if import_key in self.existing_expense_keys:
                    self.stats.expenses_skipped += 1
                    self._record_skip(
                        source_file=MANAGEMENT_FILE,
                        sheet=sheet_label,
                        row_number=row_number,
                        reason="duplicate_expense",
                        prop_id=prop.client_prop_id,
                        transaction_date=tx_date,
                        amount=exp_amount,
                        section=section,
                        notes=notes,
                        details=f"Duplicate expense ({key_kind})",
                        import_key=import_key,
                    )
                    return
                self.db.add(
                    Expense(
                        property_id=prop.id,
                        transaction_date=tx_date,
                        amount=exp_amount,
                        currency=self.settings.default_currency,
                        category=section[:255],
                        source=exp_source,
                        payment_method=exp_method,
                        vendor_name=vendor,
                        reference=method_ref or receipt,
                        description=description,
                        notes=notes,
                        receipt_ref=receipt,
                        reconciled=reconciled,
                        paid_by_resident=paid_by_resident,
                        paid_by_company=paid_by_company,
                        paid_by_owner=paid_by_owner,
                        ledger_column=ledger_column,
                        import_key=import_key,
                        source_file=MANAGEMENT_FILE,
                    )
                )
                self.existing_expense_keys.add(import_key)
                self.stats.expenses_created += 1
                counts["expenses"] += 1
                if count_key:
                    counts[count_key] = counts.get(count_key, 0) + 1

            if amount is not None:
                add_expense(
                    key_kind="expense",
                    exp_amount=amount,
                    exp_source=source,
                    exp_method=payment_method,
                )

            # MIP column — paid by the company
            if mip_paid_amount is not None:
                add_expense(
                    key_kind="mip",
                    exp_amount=mip_paid_amount,
                    exp_source="manual_company",
                    exp_method="company_account",
                    paid_by_company=True,
                    count_key="paid_by_company",
                )

            # Resident-paid (He/She paid) — excluded from company totals
            if resident_paid_amount is not None:
                add_expense(
                    key_kind="resident",
                    exp_amount=resident_paid_amount,
                    exp_source="manual_owner",
                    exp_method="owner_personal",
                    paid_by_resident=True,
                    count_key="resident_paid",
                )

            # Owner-paid (e.g. "אהרון שילם") — excluded from company totals
            if owner_paid_amount is not None:
                add_expense(
                    key_kind="owner",
                    exp_amount=owner_paid_amount,
                    exp_source="manual_owner",
                    exp_method="owner_personal",
                    paid_by_owner=True,
                    ledger_column="owner_paid",
                    count_key="owner_paid",
                )

            # Nearly CC / Cash / Other — alternate payment lanes on some property sheets
            if nearly_cc_amount is not None:
                add_expense(
                    key_kind="nearlycc",
                    exp_amount=nearly_cc_amount,
                    exp_source="credit_card",
                    exp_method="credit_card",
                    ledger_column="nearly_cc",
                    count_key="nearly_cc",
                )

            if cash_amount is not None:
                add_expense(
                    key_kind="cash",
                    exp_amount=cash_amount,
                    exp_source="management_ledger",
                    exp_method="cash",
                    ledger_column="cash",
                    count_key="cash",
                )

            if other_amount is not None:
                add_expense(
                    key_kind="other",
                    exp_amount=other_amount,
                    exp_source="management_ledger",
                    exp_method="company_account",
                    ledger_column="other",
                    count_key="other",
                )

            if inflow is not None:
                import_key = f"mgmt:{sheet_slug}:r{row_number}:inflow:{tx_date.isoformat()}:{inflow}"
                if import_key not in self.existing_deposit_keys:
                    self.db.add(
                        Deposit(
                            bank_account_id=None,
                            property_id=prop.id,
                            transaction_date=tx_date,
                            amount=inflow,
                            currency=self.settings.default_currency,
                            reference=receipt or method_ref,
                            description=section if section else "Owner inflow / prepaid",
                            source="management_ledger",
                            is_rental_income=False,
                            import_key=import_key,
                            source_file=MANAGEMENT_FILE,
                        )
                    )
                    self.existing_deposit_keys.add(import_key)
                    self.stats.deposits_created += 1
                    counts["deposits"] += 1
                else:
                    self.stats.deposits_skipped += 1
                    self._record_skip(
                        source_file=MANAGEMENT_FILE,
                        sheet=sheet_label,
                        row_number=row_number,
                        reason="duplicate_deposit",
                        prop_id=prop.client_prop_id,
                        transaction_date=tx_date,
                        inflow=inflow,
                        section=section,
                        notes=notes,
                        details="Duplicate inflow deposit",
                        import_key=import_key,
                    )

            # Rental income — shown in UI, excluded from company float totals
            if rental_income_amount is not None:
                import_key = (
                    f"mgmt:{sheet_slug}:r{row_number}:rental:"
                    f"{tx_date.isoformat()}:{rental_income_amount}"
                )
                if import_key not in self.existing_deposit_keys:
                    self.db.add(
                        Deposit(
                            bank_account_id=None,
                            property_id=prop.id,
                            transaction_date=tx_date,
                            amount=rental_income_amount,
                            currency=self.settings.default_currency,
                            reference=receipt or method_ref,
                            description=description if desc_parts else "Rental income",
                            source="rental_income",
                            is_rental_income=True,
                            import_key=import_key,
                            source_file=MANAGEMENT_FILE,
                        )
                    )
                    self.existing_deposit_keys.add(import_key)
                    self.stats.deposits_created += 1
                    counts["deposits"] += 1
                    counts["rental_income"] = counts.get("rental_income", 0) + 1
                else:
                    self.stats.deposits_skipped += 1
                    self._record_skip(
                        source_file=MANAGEMENT_FILE,
                        sheet=sheet_label,
                        row_number=row_number,
                        reason="duplicate_deposit",
                        prop_id=prop.client_prop_id,
                        transaction_date=tx_date,
                        amount=rental_income_amount,
                        section=section,
                        notes=notes,
                        details="Duplicate rental income deposit",
                        import_key=import_key,
                    )

        self.stats.sheet_counts[sheet_label] = counts

    def _detect_ledger_header(self, values: list[Any]) -> dict[str, int] | None:
        normalized = []
        for v in values:
            if v is None:
                normalized.append("")
            else:
                normalized.append(re.sub(r"\s+", " ", str(v).strip().lower()))

        joined = " | ".join(normalized)
        # Must look like a header row
        if "date" not in joined:
            return None
        if "amount" not in joined and "inflow" not in joined:
            return None

        mapping: dict[str, int] = {}
        for idx, label in enumerate(normalized):
            if not label:
                continue
            if "prop" in label and "id" in label:
                mapping["prop_id"] = idx
            elif label == "date":
                # Prefer the first Date column (ledger). Some sheets (e.g. TS 225)
                # repeat "Date" later for task tracking — do not overwrite.
                if "date" not in mapping:
                    mapping["date"] = idx
            elif label in {"section"}:
                mapping["section"] = idx
            elif label in {"notes", "note"}:
                mapping["notes"] = idx
            elif label == "type":
                mapping["type"] = idx
            elif label == "amount":
                mapping["amount"] = idx
            elif label == "inflow":
                mapping["inflow"] = idx
            elif (
                "he/she" in label
                or label in {"he/she paid", "she paid", "he paid"}
                or label.endswith(" she paid")
                or label.endswith(" he paid")
            ):
                mapping["he_she_paid"] = idx
            elif "אהרון" in label or "שילם" in label:
                mapping["owner_paid"] = idx
            elif label == "mip":
                mapping["mip"] = idx
            elif "nealy" in label or "nearly" in label:
                mapping["nearly_cc"] = idx
            elif label == "cash":
                mapping["cash"] = idx
            elif label == "other":
                mapping["other"] = idx
            elif "rental" in label or label == "rent":
                mapping["rental_income"] = idx
            elif label in {"method"}:
                mapping["method"] = idx
            elif "reconcil" in label:
                mapping["reconciled"] = idx
            elif label in {"company"}:
                mapping["company"] = idx
            elif "reciept" in label or "receipt" in label:
                mapping["receipt"] = idx

        if "date" in mapping and ("amount" in mapping or "inflow" in mapping):
            return mapping
        return None

    def _import_bank_statement(self) -> None:
        path = self.data_dir / BANK_FILE
        if not path.exists():
            self.stats.warnings.append(f"Bank statement file not found: {path.name}")
            return

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        account = self.db.scalars(
            select(BankAccount).where(BankAccount.account_number == COMPANY_ACCOUNT_NUMBER)
        ).first()
        buffer = self.properties_by_id[BUFFER_PROP_ID]

        header_row = None
        headers: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            vals = [str(v).strip() if v is not None else "" for v in row]
            if "תאריך" in vals and ("בחובה" in vals or "בזכות" in vals):
                header_row = i
                headers = vals
                break

        if header_row is None:
            self.stats.errors.append("Could not find bank statement header row")
            return

        col = {name: idx for idx, name in enumerate(headers) if name}

        def get(row_vals: list[Any], name: str) -> Any:
            idx = col.get(name)
            if idx is None or idx >= len(row_vals):
                return None
            return row_vals[idx]

        for row_number, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_number <= header_row:
                continue
            values = list(row)
            tx_date = _parse_date(get(values, "תאריך"))
            if tx_date is None:
                continue

            debit = _parse_amount(get(values, "בחובה"))
            credit = _parse_amount(get(values, "בזכות"))
            ref = _optional_str(get(values, "אסמכתא"))
            desc = _optional_str(get(values, "תיאור"))
            extended = _optional_str(get(values, "תאור מורחב"))
            full_desc = " | ".join(p for p in (desc, extended) if p)

            prop = self._guess_property_from_text(full_desc) or buffer

            if credit is not None:
                import_key = f"bank:{COMPANY_ACCOUNT_NUMBER}:r{row_number}:credit:{tx_date}:{credit}:{ref or ''}"
                if import_key not in self.existing_deposit_keys:
                    self.db.add(
                        Deposit(
                            bank_account_id=account.id if account else None,
                            property_id=prop.id,
                            transaction_date=tx_date,
                            amount=credit,
                            currency=self.settings.default_currency,
                            reference=ref,
                            description=full_desc,
                            source="bank_statement",
                            import_key=import_key,
                            source_file=BANK_FILE,
                        )
                    )
                    self.existing_deposit_keys.add(import_key)
                    self.stats.deposits_created += 1
                else:
                    self.stats.deposits_skipped += 1
                    self._record_skip(
                        source_file=BANK_FILE,
                        sheet=ws.title,
                        row_number=row_number,
                        reason="duplicate_deposit",
                        prop_id=prop.client_prop_id,
                        transaction_date=tx_date,
                        amount=credit,
                        notes=full_desc,
                        details="Duplicate bank credit",
                        import_key=import_key,
                    )

            if debit is not None:
                import_key = f"bank:{COMPANY_ACCOUNT_NUMBER}:r{row_number}:debit:{tx_date}:{debit}:{ref or ''}"
                if import_key not in self.existing_expense_keys:
                    self.db.add(
                        Expense(
                            property_id=prop.id,
                            transaction_date=tx_date,
                            amount=debit,
                            currency=self.settings.default_currency,
                            category=desc or "bank_transfer",
                            source="bank_statement",
                            payment_method="bank_transfer",
                            vendor_name=None,
                            reference=ref,
                            description=full_desc,
                            receipt_ref=ref,
                            reconciled=False,
                            import_key=import_key,
                            source_file=BANK_FILE,
                        )
                    )
                    self.existing_expense_keys.add(import_key)
                    self.stats.expenses_created += 1
                else:
                    self.stats.expenses_skipped += 1
                    self._record_skip(
                        source_file=BANK_FILE,
                        sheet=ws.title,
                        row_number=row_number,
                        reason="duplicate_expense",
                        prop_id=prop.client_prop_id,
                        transaction_date=tx_date,
                        amount=debit,
                        notes=full_desc,
                        details="Duplicate bank debit",
                        import_key=import_key,
                    )

        self.db.flush()

    def _guess_property_from_text(self, text: str | None) -> Property | None:
        if not text:
            return None
        # Match known addresses against property records
        lowered = text.lower()
        for prop in self.properties_by_id.values():
            if prop.client_prop_id == BUFFER_PROP_ID:
                continue
            if prop.address:
                addr = prop.address.strip()
                if addr and addr.lower() in lowered:
                    return prop
                # Hebrew street-number fragments often appear without city
                # e.g. address "צה\"ל 5" vs text containing צהל 5
                compact_addr = re.sub(r"\s+", "", addr).lower().replace('"', "")
                compact_text = re.sub(r"\s+", "", text).lower().replace('"', "")
                if compact_addr and compact_addr in compact_text:
                    return prop
        return None

    def _import_credit_cards(self) -> None:
        buffer = self.properties_by_id[BUFFER_PROP_ID]

        for filename in CREDIT_CARD_FILES:
            path = self.data_dir / filename
            if not path.exists():
                self.stats.warnings.append(f"Credit card file not found: {filename}")
                continue

            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb[wb.sheetnames[0]]

            card_last4 = "unknown"
            for row in ws.iter_rows(values_only=True, max_row=10):
                for cell in row:
                    if cell is None:
                        continue
                    m = re.search(r"(\d{4})\s*$", str(cell))
                    if m and "מסטרקארד" in str(cell):
                        card_last4 = m.group(1)
                        break

            account_number = f"{COMPANY_CC_ACCOUNT_PREFIX}{card_last4}"
            account = self.db.scalars(
                select(BankAccount).where(BankAccount.account_number == account_number)
            ).first()
            if not account:
                account = BankAccount(
                    property_id=None,
                    bank_name="Bank Leumi Mastercard",
                    account_number=account_number,
                    currency=self.settings.default_currency,
                    label=f"Credit card ••{card_last4}",
                )
                self.db.add(account)
                self.db.flush()
                self.stats.bank_accounts_created += 1

            header_row = None
            headers: list[str] = []
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                vals = [str(v).strip() if v is not None else "" for v in row]
                if "תאריך העסקה" in vals and "סכום חיוב" in vals:
                    header_row = i
                    headers = vals
                    break
            if header_row is None:
                self.stats.errors.append(f"No credit-card header in {filename}")
                continue

            col = {name: idx for idx, name in enumerate(headers) if name}

            def get(row_vals: list[Any], name: str) -> Any:
                idx = col.get(name)
                if idx is None or idx >= len(row_vals):
                    return None
                return row_vals[idx]

            for row_number, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_number <= header_row:
                    continue
                values = list(row)
                # Skip totals
                if any(v is not None and "סה" in str(v) for v in values[:5]):
                    continue

                tx_date = _parse_date(get(values, "תאריך העסקה"))
                merchant = _optional_str(get(values, "שם בית העסק"))
                charge = get(values, "סכום חיוב")
                # Negative charge = fee credit/refund — skip or treat as deposit
                if charge is None:
                    continue
                try:
                    charge_dec = Decimal(str(charge))
                except (InvalidOperation, ValueError):
                    continue

                if tx_date is None or merchant is None:
                    continue

                if charge_dec < 0:
                    # Card credit / fee waiver
                    amount = abs(charge_dec).quantize(Decimal("0.01"))
                    import_key = f"cc:{card_last4}:r{row_number}:credit:{tx_date}:{amount}"
                    if import_key not in self.existing_deposit_keys:
                        self.db.add(
                            Deposit(
                                bank_account_id=account.id,
                                property_id=buffer.id,
                                transaction_date=tx_date,
                                amount=amount,
                                currency=self.settings.default_currency,
                                reference=None,
                                description=merchant,
                                source="credit_card",
                                import_key=import_key,
                                source_file=filename,
                            )
                        )
                        self.existing_deposit_keys.add(import_key)
                        self.stats.deposits_created += 1
                    else:
                        self.stats.deposits_skipped += 1
                        self._record_skip(
                            source_file=filename,
                            sheet=ws.title,
                            row_number=row_number,
                            reason="duplicate_deposit",
                            transaction_date=tx_date,
                            amount=amount,
                            notes=merchant,
                            details="Duplicate credit-card credit",
                            import_key=import_key,
                        )
                    continue

                amount = charge_dec.quantize(Decimal("0.01"))
                if amount <= 0:
                    continue
                import_key = f"cc:{card_last4}:r{row_number}:expense:{tx_date}:{amount}:{merchant}"
                if import_key not in self.existing_expense_keys:
                    self.db.add(
                        Expense(
                            property_id=buffer.id,
                            transaction_date=tx_date,
                            amount=amount,
                            currency=self.settings.default_currency,
                            category=merchant[:255],
                            source="credit_card",
                            payment_method="credit_card",
                            vendor_name=merchant,
                            description=merchant,
                            import_key=import_key,
                            source_file=filename,
                        )
                    )
                    self.existing_expense_keys.add(import_key)
                    self.stats.expenses_created += 1
                else:
                    self.stats.expenses_skipped += 1
                    self._record_skip(
                        source_file=filename,
                        sheet=ws.title,
                        row_number=row_number,
                        reason="duplicate_expense",
                        transaction_date=tx_date,
                        amount=amount,
                        notes=merchant,
                        details="Duplicate credit-card expense",
                        import_key=import_key,
                    )

        self.db.flush()


def import_client_data(db: Session, data_dir: Path | None = None) -> ImportStats:
    service = ClientDataImportService(db, data_dir=data_dir)
    return service.import_all()


def build_skip_report_excel(stats: ImportStats) -> bytes:
    """Build a detailed Excel workbook of skipped import rows."""
    from collections import Counter
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()

    # Summary sheet
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Metric", "Value"])
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)
    summary_rows = [
        ("rows_seen", stats.rows_seen),
        ("skipped_rows_detailed", len(stats.skipped_rows)),
        ("rows_skipped_empty", stats.rows_skipped_empty),
        ("expenses_created", stats.expenses_created),
        ("expenses_skipped_duplicates", stats.expenses_skipped),
        ("deposits_created", stats.deposits_created),
        ("deposits_skipped_duplicates", stats.deposits_skipped),
        ("owners_created", stats.owners_created),
        ("properties_created", stats.properties_created),
        ("warnings", len(stats.warnings)),
        ("errors", len(stats.errors)),
    ]
    for key, value in summary_rows:
        summary.append([key, value])

    summary.append([])
    summary.append(["Skip reason", "Count"])
    summary[f"A{summary.max_row}"].font = Font(bold=True)
    summary[f"B{summary.max_row}"].font = Font(bold=True)
    reason_counts = Counter(row.reason for row in stats.skipped_rows)
    for reason, count in sorted(reason_counts.items(), key=lambda item: (-item[1], item[0])):
        summary.append([reason, count])

    summary.append([])
    summary.append(["Sheet", "Skipped count"])
    summary[f"A{summary.max_row}"].font = Font(bold=True)
    summary[f"B{summary.max_row}"].font = Font(bold=True)
    sheet_counts = Counter(row.sheet for row in stats.skipped_rows)
    for sheet_name, count in sorted(sheet_counts.items(), key=lambda item: (-item[1], item[0])):
        summary.append([sheet_name, count])

    # Detail sheet
    detail = wb.create_sheet("Skipped rows")
    headers = [
        "source_file",
        "sheet",
        "row_number",
        "reason",
        "prop_id",
        "transaction_date",
        "amount",
        "inflow",
        "section",
        "notes",
        "details",
        "import_key",
    ]
    detail.append(headers)
    for cell in detail[1]:
        cell.font = Font(bold=True)

    for row in stats.skipped_rows:
        data = row.to_dict()
        detail.append([data.get(h) for h in headers])

    # Reason legend
    legend = wb.create_sheet("Reason legend")
    legend.append(["reason", "meaning"])
    legend["A1"].font = Font(bold=True)
    legend["B1"].font = Font(bold=True)
    for reason, meaning in [
        ("unresolved_property", "Prop ID on the row could not be matched to a property"),
        ("missing_date", "Row has no usable transaction date"),
        ("starting_balance", "Starting-balance / opening row (intentionally skipped)"),
        ("no_money_columns", "Date present but no Amount/Inflow/paid amounts"),
        ("empty_prop_id", "Client-list row without a property number"),
        ("duplicate_expense", "Expense already exists (same import_key)"),
        ("duplicate_deposit", "Deposit already exists (same import_key)"),
    ]:
        legend.append([reason, meaning])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
