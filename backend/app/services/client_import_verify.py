"""Verify imported DB rows against client Excel source files."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import func, not_, select
from sqlalchemy.orm import Session

from app.models.deposit import Deposit
from app.models.expense import Expense
from app.models.owner import Owner
from app.models.property import Property
from app.services.client_import import (
    BANK_FILE,
    BUFFER_PROP_ID,
    CLIENT_LIST_FILE,
    CREDIT_CARD_FILES,
    MANAGEMENT_FILE,
    META_SHEETS,
    _parse_amount,
    _parse_date,
    normalize_prop_key,
    prop_key_aliases,
)


def _detect_header(values: list[Any]) -> dict[str, int] | None:
    """Match ClientDataImportService._detect_ledger_header exactly."""
    normalized = []
    for v in values:
        if v is None:
            normalized.append("")
        else:
            normalized.append(re.sub(r"\s+", " ", str(v).strip().lower()))

    joined = " | ".join(normalized)
    if "date" not in joined:
        return None
    if "amount" not in joined and "inflow" not in joined:
        return None

    mapping: dict[str, int] = {}
    for idx, label in enumerate(normalized):
        if not label:
            continue
        if "prop" in label and "id" in label:
            mapping["prop_id"] = idx
        elif label == "date":
            # Prefer the first Date column (ledger); ignore later task-tracker Date cols
            if "date" not in mapping:
                mapping["date"] = idx
        elif label in {"section"}:
            mapping["section"] = idx
        elif label in {"notes", "note"}:
            mapping["notes"] = idx
        elif label == "type":
            mapping["type"] = idx
        elif label == "amount":
            mapping["amount"] = idx
        elif label == "inflow":
            mapping["inflow"] = idx
        elif (
            "he/she" in label
            or label in {"he/she paid", "she paid", "he paid"}
            or label.endswith(" she paid")
            or label.endswith(" he paid")
        ):
            mapping["he_she_paid"] = idx
        elif "אהרון" in label or "שילם" in label:
            mapping["owner_paid"] = idx
        elif label == "mip":
            mapping["mip"] = idx
        elif "nealy" in label or "nearly" in label:
            mapping["nearly_cc"] = idx
        elif label == "cash":
            mapping["cash"] = idx
        elif label == "other":
            mapping["other"] = idx
        elif "rental" in label or label == "rent":
            mapping["rental_income"] = idx
        elif label in {"method"}:
            mapping["method"] = idx
        elif "reconcil" in label:
            mapping["reconciled"] = idx
        elif label in {"company"}:
            mapping["company"] = idx
        elif "reciept" in label or "receipt" in label:
            mapping["receipt"] = idx

    if "date" in mapping and ("amount" in mapping or "inflow" in mapping):
        return mapping
    return None


def _sheet_default_prop_id(sheet_name: str) -> str | None:
    lower = sheet_name.strip().lower()
    if lower == "buffer":
        return BUFFER_PROP_ID
    if sheet_name.strip() in {"801-618- 619", "801-618-619"}:
        return None
    return normalize_prop_key(sheet_name)


def _build_prop_alias_map(data_dir: Path) -> dict[str, str]:
    """Map normalized aliases → canonical client_prop_id (mirrors importer)."""
    alias_to_prop: dict[str, str] = {BUFFER_PROP_ID: BUFFER_PROP_ID}
    client_list = data_dir / CLIENT_LIST_FILE
    if not client_list.exists():
        return alias_to_prop
    wb = openpyxl.load_workbook(client_list, data_only=True)
    try:
        for sheet_name in ("current clients", "past clients"):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                if i == 1:
                    continue
                raw = row[0] if row else None
                canonical = normalize_prop_key(raw)
                if not canonical:
                    continue
                # Prefer first (current) registration; don't overwrite with past
                if canonical not in alias_to_prop:
                    alias_to_prop[canonical] = canonical
                for alias in prop_key_aliases(raw):
                    alias_to_prop.setdefault(alias, canonical)
    finally:
        wb.close()
    return alias_to_prop


def _resolve_verify_prop_id(
    prop_raw: Any,
    default_prop_id: str | None,
    alias_to_prop: dict[str, str],
) -> str | None:
    """Match ClientDataImportService ledger property resolution order."""
    resolved: str | None = None
    if prop_raw is not None:
        text = str(prop_raw).strip()
        # Slash/plus codes do not normalize to a single property (importer returns None)
        if text and "/" not in text and "+" not in text:
            key = normalize_prop_key(prop_raw)
            if key:
                if key in alias_to_prop:
                    resolved = alias_to_prop[key]
                else:
                    for alias in prop_key_aliases(prop_raw):
                        if alias in alias_to_prop:
                            resolved = alias_to_prop[alias]
                            break
    if resolved is None and default_prop_id:
        resolved = alias_to_prop.get(default_prop_id, default_prop_id)
    if resolved is None and prop_raw is not None:
        text = str(prop_raw)
        if "/" in text or "+" in text:
            return BUFFER_PROP_ID
    return resolved


def count_management_rows(path: Path) -> dict[str, int]:
    """Count unique management ledger import keys (matches importer dedupe)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    alias_to_prop = _build_prop_alias_map(path.parent)
    expense_keys: set[str] = set()
    deposit_keys: set[str] = set()
    resident_keys: set[str] = set()
    owner_keys: set[str] = set()
    mip_keys: set[str] = set()
    nearly_keys: set[str] = set()
    cash_keys: set[str] = set()
    other_keys: set[str] = set()
    rental_keys: set[str] = set()
    sheets_parsed = 0

    # Register sheet-name aliases like the importer does when creating ledger props
    for sheet_name in wb.sheetnames:
        lower = sheet_name.strip().lower()
        if lower in META_SHEETS and lower != "buffer":
            continue
        if sheet_name.strip() in {"801-618- 619", "801-618-619"}:
            continue
        if lower == "buffer":
            continue
        canonical = normalize_prop_key(sheet_name)
        if not canonical:
            continue
        alias_to_prop.setdefault(canonical, canonical)
        for alias in prop_key_aliases(sheet_name):
            alias_to_prop.setdefault(alias, canonical)

    for sheet_name in wb.sheetnames:
        lower = sheet_name.strip().lower()
        # Skip meta sheets except Buffer (company float ledger)
        if lower in META_SHEETS and lower != "buffer":
            continue

        default_prop_id = _sheet_default_prop_id(sheet_name)
        if default_prop_id:
            default_prop_id = alias_to_prop.get(default_prop_id, default_prop_id)
        sheet_slug = re.sub(r"[^A-Za-z0-9]+", "", sheet_name).lower() or "sheet"
        ws = wb[sheet_name]
        header = None
        for row in ws.iter_rows(values_only=True):
            values = list(row)
            if header is None:
                detected = _detect_header(values)
                if detected:
                    header = detected
                continue
            if not any(v is not None and str(v).strip() for v in values):
                continue

            def col(name: str) -> Any:
                idx = header.get(name)
                if idx is None or idx >= len(values):
                    return None
                return values[idx]

            prop_id = _resolve_verify_prop_id(col("prop_id"), default_prop_id, alias_to_prop)
            if prop_id is None:
                continue

            tx_date = _parse_date(col("date"))
            amount = _parse_amount(col("amount"))
            inflow = _parse_amount(col("inflow"))
            resident = _parse_amount(col("he_she_paid"))
            owner = _parse_amount(col("owner_paid"))
            mip = _parse_amount(col("mip"))
            nearly_cc = _parse_amount(col("nearly_cc"))
            cash = _parse_amount(col("cash"))
            other = _parse_amount(col("other"))
            rental = _parse_amount(col("rental_income"))
            if tx_date is None:
                continue

            section_raw = col("section") if col("section") is not None else col("type")
            section = str(section_raw).strip() if section_raw is not None else ""
            section = (section or "other")[:40]
            date_key = tx_date.isoformat()

            def add_expense(kind: str, exp_amount: Any, bucket: set[str] | None = None) -> None:
                key = (
                    f"mgmt:{sheet_slug}:{kind}:{date_key}:{exp_amount}:"
                    f"{prop_id}:{section}"
                )
                expense_keys.add(key)
                if bucket is not None:
                    bucket.add(key)

            def add_deposit(kind: str, dep_amount: Any, bucket: set[str] | None = None) -> None:
                key = (
                    f"mgmt:{sheet_slug}:{kind}:{date_key}:{dep_amount}:"
                    f"{prop_id}:{section}"
                )
                deposit_keys.add(key)
                if bucket is not None:
                    bucket.add(key)

            if amount is not None:
                add_expense("expense", amount)
            if resident is not None:
                add_expense("resident", resident, resident_keys)
            if owner is not None:
                add_expense("owner", owner, owner_keys)
            if mip is not None:
                add_expense("mip", mip, mip_keys)
            if nearly_cc is not None:
                add_expense("nearlycc", nearly_cc, nearly_keys)
            if cash is not None:
                add_expense("cash", cash, cash_keys)
            if other is not None:
                add_expense("other", other, other_keys)
            if inflow is not None:
                add_deposit("inflow", inflow)
            if rental is not None:
                add_deposit("rental", rental, rental_keys)
        if header is not None:
            sheets_parsed += 1

    return {
        "mgmt_expense_rows": len(expense_keys),
        "mgmt_deposit_rows": len(deposit_keys),
        "mgmt_resident_paid_rows": len(resident_keys),
        "mgmt_owner_paid_rows": len(owner_keys),
        "mgmt_mip_paid_rows": len(mip_keys),
        "mgmt_nearly_cc_rows": len(nearly_keys),
        "mgmt_cash_rows": len(cash_keys),
        "mgmt_other_rows": len(other_keys),
        "mgmt_rental_income_rows": len(rental_keys),
        "sheets_parsed": sheets_parsed,
    }


def count_bank_rows(path: Path) -> dict[str, int]:
    if not path.exists():
        return {"bank_debit_rows": 0, "bank_credit_rows": 0}
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = None
    headers: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        vals = [str(v).strip() if v is not None else "" for v in row]
        if "תאריך" in vals and ("בחובה" in vals or "בזכות" in vals):
            header_row = i
            headers = vals
            break
    if header_row is None:
        return {"bank_debit_rows": 0, "bank_credit_rows": 0}

    col = {name: idx for idx, name in enumerate(headers) if name}
    debits = 0
    credits = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= header_row:
            continue
        values = list(row)
        date_idx = col.get("תאריך")
        if date_idx is None or date_idx >= len(values) or _parse_date(values[date_idx]) is None:
            continue
        debit_idx = col.get("בחובה")
        credit_idx = col.get("בזכות")
        if debit_idx is not None and debit_idx < len(values) and _parse_amount(values[debit_idx]):
            debits += 1
        if credit_idx is not None and credit_idx < len(values) and _parse_amount(values[credit_idx]):
            credits += 1
    return {"bank_debit_rows": debits, "bank_credit_rows": credits}


def count_cc_rows(data_dir: Path) -> dict[str, int]:
    total = 0
    credits = 0
    for name in CREDIT_CARD_FILES:
        path = data_dir / name
        if not path.exists():
            continue
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        header_row = None
        headers: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            vals = [str(v).strip() if v is not None else "" for v in row]
            if "תאריך העסקה" in vals and "סכום חיוב" in vals:
                header_row = i
                headers = vals
                break
        if header_row is None:
            continue
        col = {name: idx for idx, name in enumerate(headers) if name}
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i <= header_row:
                continue
            values = list(row)
            if any(v is not None and "סה" in str(v) for v in values[:5]):
                continue
            date_idx = col.get("תאריך העסקה")
            charge_idx = col.get("סכום חיוב")
            merchant_idx = col.get("שם בית העסק")
            if date_idx is None or charge_idx is None:
                continue
            if date_idx >= len(values) or _parse_date(values[date_idx]) is None:
                continue
            if merchant_idx is not None and merchant_idx < len(values):
                if values[merchant_idx] is None:
                    continue
            try:
                charge = Decimal(str(values[charge_idx]))
            except Exception:
                continue
            if charge < 0:
                credits += 1
            elif charge > 0:
                total += 1
    return {"cc_expense_rows": total, "cc_credit_rows": credits}


def count_client_list_properties(path: Path) -> int:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["current clients"]
    keys: set[str] = set()
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue
        key = normalize_prop_key(row[0] if row else None)
        if key:
            keys.add(key)
    return len(keys)


def verify_against_excel(db: Session, data_dir: Path) -> dict[str, Any]:
    mgmt = count_management_rows(data_dir / MANAGEMENT_FILE)
    bank = count_bank_rows(data_dir / BANK_FILE)
    cc = count_cc_rows(data_dir)
    client_props = count_client_list_properties(data_dir / CLIENT_LIST_FILE)

    expected_expenses = (
        mgmt["mgmt_expense_rows"] + bank["bank_debit_rows"] + cc["cc_expense_rows"]
    )
    expected_deposits = (
        mgmt["mgmt_deposit_rows"] + bank["bank_credit_rows"] + cc["cc_credit_rows"]
    )

    db_expenses = db.scalar(select(func.count()).select_from(Expense)) or 0
    db_deposits = db.scalar(select(func.count()).select_from(Deposit)) or 0
    db_props = db.scalar(select(func.count()).select_from(Property)) or 0
    db_owners = db.scalar(select(func.count()).select_from(Owner)) or 0

    ledger_expenses = (
        db.scalar(
            select(func.count()).select_from(Expense).where(Expense.source == "management_ledger")
        )
        or 0
    )
    # Complete mgmt rows only — incomplete/needs_review keys are not in Excel unique counts
    mgmt_complete = not_(Expense.import_key.like("%:incomplete:%"))
    mgmt_dep_complete = not_(Deposit.import_key.like("%:incomplete:%"))
    mgmt_exp_db = (
        db.scalar(
            select(func.count())
            .select_from(Expense)
            .where(Expense.import_key.like("mgmt:%"), mgmt_complete)
        )
        or 0
    )
    mgmt_dep_db = (
        db.scalar(
            select(func.count())
            .select_from(Deposit)
            .where(Deposit.import_key.like("mgmt:%"), mgmt_dep_complete)
        )
        or 0
    )

    mgmt_resident_db = (
        db.scalar(
            select(func.count())
            .select_from(Expense)
            .where(Expense.paid_by_resident.is_(True), mgmt_complete)
        )
        or 0
    )

    mgmt_rental_db = (
        db.scalar(
            select(func.count())
            .select_from(Deposit)
            .where(Deposit.is_rental_income.is_(True), mgmt_dep_complete)
        )
        or 0
    )

    mgmt_mip_db = (
        db.scalar(
            select(func.count())
            .select_from(Expense)
            .where(Expense.paid_by_company.is_(True), mgmt_complete)
        )
        or 0
    )

    mgmt_owner_db = (
        db.scalar(
            select(func.count())
            .select_from(Expense)
            .where(Expense.paid_by_owner.is_(True), mgmt_complete)
        )
        or 0
    )

    mgmt_nearly_db = (
        db.scalar(
            select(func.count())
            .select_from(Expense)
            .where(Expense.ledger_column == "nearly_cc", mgmt_complete)
        )
        or 0
    )
    mgmt_cash_db = (
        db.scalar(
            select(func.count())
            .select_from(Expense)
            .where(Expense.ledger_column == "cash", mgmt_complete)
        )
        or 0
    )
    mgmt_other_db = (
        db.scalar(
            select(func.count())
            .select_from(Expense)
            .where(Expense.ledger_column == "other", mgmt_complete)
        )
        or 0
    )

    mismatches: list[str] = []
    if mgmt_exp_db != mgmt["mgmt_expense_rows"]:
        mismatches.append(
            f"Management expenses: excel={mgmt['mgmt_expense_rows']} db={mgmt_exp_db}"
        )
    if mgmt_dep_db != mgmt["mgmt_deposit_rows"]:
        mismatches.append(
            f"Management deposits: excel={mgmt['mgmt_deposit_rows']} db={mgmt_dep_db}"
        )
    if mgmt_resident_db != mgmt["mgmt_resident_paid_rows"]:
        mismatches.append(
            f"Resident-paid expenses: excel={mgmt['mgmt_resident_paid_rows']} db={mgmt_resident_db}"
        )
    if mgmt_owner_db != mgmt["mgmt_owner_paid_rows"]:
        mismatches.append(
            f"Owner-paid expenses: excel={mgmt['mgmt_owner_paid_rows']} db={mgmt_owner_db}"
        )
    if mgmt_mip_db != mgmt["mgmt_mip_paid_rows"]:
        mismatches.append(
            f"MIP-paid expenses: excel={mgmt['mgmt_mip_paid_rows']} db={mgmt_mip_db}"
        )
    if mgmt_nearly_db != mgmt["mgmt_nearly_cc_rows"]:
        mismatches.append(
            f"Nearly CC expenses: excel={mgmt['mgmt_nearly_cc_rows']} db={mgmt_nearly_db}"
        )
    if mgmt_cash_db != mgmt["mgmt_cash_rows"]:
        mismatches.append(
            f"Cash expenses: excel={mgmt['mgmt_cash_rows']} db={mgmt_cash_db}"
        )
    if mgmt_other_db != mgmt["mgmt_other_rows"]:
        mismatches.append(
            f"Other-column expenses: excel={mgmt['mgmt_other_rows']} db={mgmt_other_db}"
        )
    if mgmt_rental_db != mgmt["mgmt_rental_income_rows"]:
        mismatches.append(
            f"Rental income: excel={mgmt['mgmt_rental_income_rows']} db={mgmt_rental_db}"
        )

    # Properties: at least current client list + BUFFER
    if db_props < client_props + 1:
        mismatches.append(
            f"Properties: expected at least {client_props + 1} (clients+BUFFER), db={db_props}"
        )

    ok = len(mismatches) == 0
    return {
        "ok": ok,
        "mismatches": mismatches,
        "excel": {
            "current_client_properties": client_props,
            **mgmt,
            **bank,
            **cc,
            "expected_expenses_total": expected_expenses,
            "expected_deposits_total": expected_deposits,
        },
        "database": {
            "owners": db_owners,
            "properties": db_props,
            "expenses": db_expenses,
            "deposits": db_deposits,
            "mgmt_expenses": mgmt_exp_db,
            "mgmt_deposits": mgmt_dep_db,
            "resident_paid_expenses": mgmt_resident_db,
            "owner_paid_expenses": mgmt_owner_db,
            "mip_paid_expenses": mgmt_mip_db,
            "nearly_cc_expenses": mgmt_nearly_db,
            "cash_expenses": mgmt_cash_db,
            "other_column_expenses": mgmt_other_db,
            "rental_income_deposits": mgmt_rental_db,
            "ledger_source_expenses": ledger_expenses,
        },
    }
