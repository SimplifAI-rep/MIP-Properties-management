"""Parse bank-statement and credit-card Excel uploads into review drafts.

Creates link-or-create style drafts with duplicate detection. User confirms
each row as add or ignore — existing transactions are never updated.
"""

from __future__ import annotations

import json
import logging
import re
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Literal
from uuid import UUID

import httpx
import openpyxl
from sqlalchemy import and_, or_, select
from sqlalchemy.orm import Session, joinedload

from app.core.config import get_settings
from app.models.bank_account import BankAccount
from app.models.deposit import Deposit
from app.models.expense import Expense
from app.models.owner import Owner
from app.models.property import Property
from app.schemas import FieldWarning, TransactionDraft

logger = logging.getLogger(__name__)

BUFFER_PROP_ID = "BUFFER"
COMPANY_OWNER_NAME = "My Israel Property (MIP)"
COMPANY_ACCOUNT_NUMBER = "MIP-LEUMI-OPS"
COMPANY_CC_ACCOUNT_PREFIX = "MIP-LEUMI-CC-"

StatementKind = Literal["bank_statement", "credit_card"]


def detect_statement_kind(content: bytes) -> StatementKind | None:
    """Return bank_statement / credit_card when Hebrew Leumi headers are present."""
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception:
        return None
    try:
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(values_only=True, max_row=40):
            vals = [str(v).strip() if v is not None else "" for v in row]
            if "תאריך העסקה" in vals and "סכום חיוב" in vals:
                return "credit_card"
            if "תאריך" in vals and ("בחובה" in vals or "בזכות" in vals):
                return "bank_statement"
    finally:
        wb.close()
    return None


def _optional_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        return None
    else:
        text = str(value).strip()
        if not text or text.lower() in {"none", "nan", "nat"}:
            return None
        parsed = None
        try:
            parsed = datetime.fromisoformat(text.replace("Z", "")).date()
        except ValueError:
            pass
        if parsed is None:
            for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d.%m.%y", "%d/%m/%y"):
                try:
                    parsed = datetime.strptime(text, fmt).date()
                    break
                except ValueError:
                    continue
        if parsed is None:
            return None
    if parsed.year < 2000:
        return None
    return parsed


def _parse_amount(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        amount = value
    else:
        text = str(value).strip().replace(",", "").replace("₪", "").replace("NIS", "")
        if not text or text.lower() in {"none", "nan"}:
            return None
        try:
            amount = Decimal(text)
        except (InvalidOperation, ValueError):
            return None
    if amount == 0:
        return None
    return abs(amount).quantize(Decimal("0.01"))


def _signed_amount(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        amount = value
    else:
        text = str(value).strip().replace(",", "").replace("₪", "").replace("NIS", "")
        if not text or text.lower() in {"none", "nan"}:
            return None
        try:
            amount = Decimal(text)
        except (InvalidOperation, ValueError):
            return None
    if amount == 0:
        return None
    return amount.quantize(Decimal("0.01"))


class StatementImportService:
    def __init__(self, db: Session):
        self.db = db
        self.settings = get_settings()

    def analyze(
        self,
        content: bytes,
        *,
        filename: str,
        kind: StatementKind,
    ) -> tuple[list[TransactionDraft], str, str | None]:
        buffer = self._ensure_buffer_property()
        if kind == "bank_statement":
            drafts = self._parse_bank(content, filename=filename, buffer=buffer)
            parser = "bank_statement_excel"
        else:
            drafts = self._parse_credit_card(content, filename=filename, buffer=buffer)
            parser = "credit_card_excel"

        if not drafts:
            empty = TransactionDraft(
                transaction_type="expense",
                property_id=buffer.id,
                client_prop_id=buffer.client_prop_id,
                property_name=buffer.name,
                owner_id=buffer.owner_id,
                owner_name=buffer.owner.name if buffer.owner else None,
                status="error",
                user_action="ignore",
                needs_review=True,
                review_reasons="no_rows",
                warnings=[
                    FieldWarning(
                        field="document",
                        message="No transaction rows found in this file.",
                        severity="error",
                    )
                ],
            )
            return [empty], parser, "No transaction rows found."

        self._attach_duplicates(drafts)
        self._enrich_with_ai(drafts, buffer=buffer)
        message = (
            f"Parsed {len(drafts)} row(s). "
            "Review each row — duplicates can be added again or ignored. "
            "Rows on the company buffer need confirmation."
        )
        return drafts, parser, message

    def _ensure_buffer_property(self) -> Property:
        buffer = self.db.scalars(
            select(Property)
            .options(joinedload(Property.owner))
            .where(Property.client_prop_id == BUFFER_PROP_ID)
        ).first()
        if buffer:
            return buffer

        owner = self.db.scalars(
            select(Owner).where(Owner.name == COMPANY_OWNER_NAME)
        ).first()
        if not owner:
            owner = Owner(name=COMPANY_OWNER_NAME)
            self.db.add(owner)
            self.db.flush()

        buffer = Property(
            owner_id=owner.id,
            client_prop_id=BUFFER_PROP_ID,
            name="MIP Company Buffer",
            address="Company float / unallocated",
            city=None,
            status="active",
        )
        self.db.add(buffer)
        self.db.flush()
        return self.db.scalars(
            select(Property)
            .options(joinedload(Property.owner))
            .where(Property.id == buffer.id)
        ).one()

    def _ensure_company_bank_account(self) -> BankAccount:
        account = self.db.scalars(
            select(BankAccount).where(BankAccount.account_number == COMPANY_ACCOUNT_NUMBER)
        ).first()
        if account:
            return account
        account = BankAccount(
            property_id=None,
            bank_name="Bank Leumi",
            account_number=COMPANY_ACCOUNT_NUMBER,
            currency=self.settings.default_currency,
            label="MIP operating account",
        )
        self.db.add(account)
        self.db.flush()
        return account

    def _ensure_cc_account(self, card_last4: str) -> BankAccount:
        account_number = f"{COMPANY_CC_ACCOUNT_PREFIX}{card_last4}"
        account = self.db.scalars(
            select(BankAccount).where(BankAccount.account_number == account_number)
        ).first()
        if account:
            return account
        account = BankAccount(
            property_id=None,
            bank_name="Bank Leumi Mastercard",
            account_number=account_number,
            currency=self.settings.default_currency,
            label=f"Credit card ••{card_last4}",
        )
        self.db.add(account)
        self.db.flush()
        return account

    def _property_meta(self, prop: Property) -> dict[str, Any]:
        return {
            "property_id": prop.id,
            "client_prop_id": prop.client_prop_id,
            "property_name": prop.name,
            "owner_id": prop.owner_id,
            "owner_name": prop.owner.name if prop.owner else None,
        }

    def _guess_property_from_text(self, text: str | None) -> Property | None:
        if not text:
            return None
        lowered = text.lower()
        props = self.db.scalars(
            select(Property).options(joinedload(Property.owner))
        ).unique().all()
        for prop in props:
            if prop.client_prop_id == BUFFER_PROP_ID:
                continue
            if prop.address:
                addr = prop.address.strip()
                if addr and addr.lower() in lowered:
                    return prop
                compact_addr = re.sub(r"\s+", "", addr).lower().replace('"', "")
                compact_text = re.sub(r"\s+", "", text).lower().replace('"', "")
                if compact_addr and compact_addr in compact_text:
                    return prop
            if prop.name and prop.name.lower() in lowered:
                return prop
            if prop.client_prop_id and prop.client_prop_id.lower() in lowered:
                return prop
        return None

    def _parse_bank(
        self,
        content: bytes,
        *,
        filename: str,
        buffer: Property,
    ) -> list[TransactionDraft]:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        drafts: list[TransactionDraft] = []
        try:
            ws = wb[wb.sheetnames[0]]
            account = self._ensure_company_bank_account()
            header_row = None
            headers: list[str] = []
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                vals = [str(v).strip() if v is not None else "" for v in row]
                if "תאריך" in vals and ("בחובה" in vals or "בזכות" in vals):
                    header_row = i
                    headers = vals
                    break
            if header_row is None:
                return []

            col = {name: idx for idx, name in enumerate(headers) if name}

            def get(row_vals: list[Any], name: str) -> Any:
                idx = col.get(name)
                if idx is None or idx >= len(row_vals):
                    return None
                return row_vals[idx]

            for row_number, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_number <= header_row:
                    continue
                values = list(row)
                tx_date = _parse_date(get(values, "תאריך"))

                debit = _parse_amount(get(values, "בחובה"))
                credit = _parse_amount(get(values, "בזכות"))
                # Treat 0 cells as empty (sample file uses 0 placeholders)
                raw_debit = get(values, "בחובה")
                raw_credit = get(values, "בזכות")
                try:
                    if raw_debit is not None and Decimal(str(raw_debit)) == 0:
                        debit = None
                except (InvalidOperation, ValueError):
                    pass
                try:
                    if raw_credit is not None and Decimal(str(raw_credit)) == 0:
                        credit = None
                except (InvalidOperation, ValueError):
                    pass

                ref = _optional_str(get(values, "אסמכתא"))
                desc = _optional_str(get(values, "תיאור"))
                extended = _optional_str(get(values, "תאור מורחב"))
                full_desc = " | ".join(p for p in (desc, extended) if p)

                # Skip totally blank rows.
                if (
                    tx_date is None
                    and debit is None
                    and credit is None
                    and not full_desc
                    and not ref
                ):
                    continue

                guessed = self._guess_property_from_text(full_desc)
                prop = guessed or buffer
                meta = self._property_meta(prop)
                on_buffer = prop.client_prop_id == BUFFER_PROP_ID

                if credit is None and debit is None:
                    drafts.append(
                        self._make_draft(
                            row_number=row_number,
                            transaction_type="expense",
                            meta=meta,
                            tx_date=tx_date,
                            amount=None,
                            description=full_desc,
                            reference=ref,
                            source="bank_statement",
                            payment_method="bank_transfer",
                            category=desc or "bank_transfer",
                            vendor_name=None,
                            bank_account_id=account.id,
                            account_number=account.account_number,
                            on_buffer=on_buffer,
                            match_confidence="low",
                            import_key=(
                                f"bank:{COMPANY_ACCOUNT_NUMBER}:incomplete:"
                                f"{tx_date or 'nodate'}:{ref or ''}:{full_desc[:80]}"
                            ),
                            incomplete_reasons=(
                                (["missing_date"] if tx_date is None else [])
                                + ["missing_amount"]
                            ),
                        )
                    )
                    continue

                if credit is not None:
                    incomplete = ["missing_date"] if tx_date is None else []
                    drafts.append(
                        self._make_draft(
                            row_number=row_number,
                            transaction_type="deposit",
                            meta=meta,
                            tx_date=tx_date,
                            amount=credit,
                            description=full_desc,
                            reference=ref,
                            source="bank_statement",
                            payment_method=None,
                            category=None,
                            vendor_name=None,
                            bank_account_id=account.id,
                            account_number=account.account_number,
                            on_buffer=on_buffer,
                            match_confidence=(
                                "medium" if guessed and not incomplete else "low"
                            ),
                            import_key=(
                                f"bank:{COMPANY_ACCOUNT_NUMBER}:credit:"
                                f"{tx_date or 'nodate'}:{credit}:{ref or ''}:"
                                f"{full_desc[:60]}"
                            ),
                            incomplete_reasons=incomplete or None,
                        )
                    )
                if debit is not None:
                    incomplete = ["missing_date"] if tx_date is None else []
                    drafts.append(
                        self._make_draft(
                            row_number=row_number,
                            transaction_type="expense",
                            meta=meta,
                            tx_date=tx_date,
                            amount=debit,
                            description=full_desc,
                            reference=ref,
                            source="bank_statement",
                            payment_method="bank_transfer",
                            category=desc or "bank_transfer",
                            vendor_name=None,
                            bank_account_id=None,
                            account_number=None,
                            on_buffer=on_buffer,
                            match_confidence=(
                                "medium" if guessed and not incomplete else "low"
                            ),
                            import_key=(
                                f"bank:{COMPANY_ACCOUNT_NUMBER}:debit:"
                                f"{tx_date or 'nodate'}:{debit}:{ref or ''}:"
                                f"{full_desc[:60]}"
                            ),
                            incomplete_reasons=incomplete or None,
                        )
                    )
        finally:
            wb.close()
        return drafts

    def _parse_credit_card(
        self,
        content: bytes,
        *,
        filename: str,
        buffer: Property,
    ) -> list[TransactionDraft]:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        drafts: list[TransactionDraft] = []
        try:
            ws = wb[wb.sheetnames[0]]
            card_last4 = "unknown"
            for row in ws.iter_rows(values_only=True, max_row=10):
                for cell in row:
                    if cell is None:
                        continue
                    text = str(cell)
                    m = re.search(r"(\d{4})\s*$", text)
                    if m and ("מסטרקארד" in text or "כרטיס" in text):
                        card_last4 = m.group(1)
                        break

            account = self._ensure_cc_account(card_last4)
            header_row = None
            headers: list[str] = []
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                vals = [str(v).strip() if v is not None else "" for v in row]
                if "תאריך העסקה" in vals and "סכום חיוב" in vals:
                    header_row = i
                    headers = vals
                    break
            if header_row is None:
                return []

            col = {name: idx for idx, name in enumerate(headers) if name}

            def get(row_vals: list[Any], name: str) -> Any:
                idx = col.get(name)
                if idx is None or idx >= len(row_vals):
                    return None
                return row_vals[idx]

            buffer_meta = self._property_meta(buffer)
            for row_number, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_number <= header_row:
                    continue
                values = list(row)
                if any(v is not None and "סה" in str(v) for v in values[:5]):
                    continue

                tx_date = _parse_date(get(values, "תאריך העסקה"))
                merchant = _optional_str(get(values, "שם בית העסק"))
                charge = _signed_amount(get(values, "סכום חיוב"))
                if charge is None and tx_date is None and merchant is None:
                    continue

                guessed = self._guess_property_from_text(merchant or "")
                prop = guessed or buffer
                meta = self._property_meta(prop)
                on_buffer = prop.client_prop_id == BUFFER_PROP_ID
                incomplete: list[str] = []
                if tx_date is None:
                    incomplete.append("missing_date")
                if charge is None:
                    incomplete.append("missing_amount")
                if merchant is None:
                    incomplete.append("missing_merchant")

                if charge is None or incomplete:
                    drafts.append(
                        self._make_draft(
                            row_number=row_number,
                            transaction_type="expense",
                            meta=meta,
                            tx_date=tx_date,
                            amount=abs(charge) if charge is not None else None,
                            description=merchant,
                            reference=None,
                            source="credit_card",
                            payment_method="credit_card",
                            category=(merchant[:255] if merchant else "credit_card"),
                            vendor_name=merchant,
                            bank_account_id=None,
                            account_number=None,
                            on_buffer=on_buffer,
                            match_confidence="low",
                            import_key=(
                                f"cc:{card_last4}:incomplete:"
                                f"{tx_date or 'nodate'}:{charge}:{merchant or ''}"
                            ),
                            incomplete_reasons=incomplete or ["missing_amount"],
                        )
                    )
                    continue

                if charge < 0:
                    amount = abs(charge)
                    drafts.append(
                        self._make_draft(
                            row_number=row_number,
                            transaction_type="deposit",
                            meta=meta if not on_buffer else buffer_meta,
                            tx_date=tx_date,
                            amount=amount,
                            description=merchant,
                            reference=None,
                            source="credit_card",
                            payment_method=None,
                            category=None,
                            vendor_name=merchant,
                            bank_account_id=account.id,
                            account_number=account.account_number,
                            on_buffer=True if not guessed else on_buffer,
                            match_confidence="medium" if guessed else "low",
                            import_key=(
                                f"cc:{card_last4}:credit:{tx_date}:{amount}:{merchant}"
                            ),
                        )
                    )
                    continue

                amount = charge
                drafts.append(
                    self._make_draft(
                        row_number=row_number,
                        transaction_type="expense",
                        meta=meta,
                        tx_date=tx_date,
                        amount=amount,
                        description=merchant,
                        reference=None,
                        source="credit_card",
                        payment_method="credit_card",
                        category=merchant[:255],
                        vendor_name=merchant,
                        bank_account_id=None,
                        account_number=None,
                        on_buffer=on_buffer,
                        match_confidence="medium" if guessed else "low",
                        import_key=(
                            f"cc:{card_last4}:expense:{tx_date}:{amount}:{merchant}"
                        ),
                    )
                )
        finally:
            wb.close()
        return drafts

    def _make_draft(
        self,
        *,
        row_number: int,
        transaction_type: Literal["deposit", "expense"],
        meta: dict[str, Any],
        tx_date: date | None,
        amount: Decimal | None,
        description: str | None,
        reference: str | None,
        source: str,
        payment_method: str | None,
        category: str | None,
        vendor_name: str | None,
        bank_account_id: UUID | None,
        account_number: str | None,
        on_buffer: bool,
        match_confidence: Literal["high", "medium", "low", "none"],
        import_key: str,
        incomplete_reasons: list[str] | None = None,
    ) -> TransactionDraft:
        warnings: list[FieldWarning] = []
        reasons = list(incomplete_reasons or [])
        if on_buffer:
            warnings.append(
                FieldWarning(
                    field="property_id",
                    message="Assigned to company buffer — confirm or change the property.",
                    severity="warning",
                )
            )
        if "missing_date" in reasons:
            warnings.append(
                FieldWarning(
                    field="transaction_date",
                    message="Missing or unreadable date — fill it in before adding.",
                    severity="error",
                )
            )
        if "missing_amount" in reasons:
            warnings.append(
                FieldWarning(
                    field="amount",
                    message="Missing or unreadable amount — fill it in before adding.",
                    severity="error",
                )
            )
        if "missing_merchant" in reasons:
            warnings.append(
                FieldWarning(
                    field="vendor_name",
                    message="Missing merchant/description — fill it in before adding.",
                    severity="warning",
                )
            )
        incomplete = bool(reasons)
        return TransactionDraft(
            row_number=row_number,
            transaction_type=transaction_type,
            property_id=meta["property_id"],
            client_prop_id=meta["client_prop_id"],
            property_name=meta["property_name"],
            owner_id=meta["owner_id"],
            owner_name=meta["owner_name"],
            bank_account_id=bank_account_id,
            account_number=account_number,
            transaction_date=tx_date,
            amount=amount,
            currency=self.settings.default_currency,
            category=category,
            source=source,
            payment_method=payment_method,
            vendor_name=vendor_name,
            reference=reference,
            description=description,
            match_confidence=match_confidence,
            status="error" if incomplete else "needs_review",
            warnings=warnings,
            user_action="ignore" if incomplete else "add",
            is_duplicate=False,
            needs_review=incomplete,
            review_reasons=",".join(reasons) if reasons else None,
            import_key=import_key,
        )

    def _attach_duplicates(self, drafts: list[TransactionDraft]) -> None:
        for draft in drafts:
            if not draft.transaction_date or draft.amount is None:
                continue
            match = self._find_duplicate(draft)
            if not match:
                continue
            kind, row = match
            draft.is_duplicate = True
            draft.duplicate_match_id = row.id
            draft.duplicate_match_kind = kind
            draft.duplicate_summary = (
                f"{kind} {row.transaction_date} · {row.amount} "
                f"· {(getattr(row, 'description', None) or getattr(row, 'category', None) or '')[:80]}"
            )
            draft.user_action = "ignore"
            draft.warnings.insert(
                0,
                FieldWarning(
                    field="duplicate",
                    message=f"Possible duplicate of existing {kind}: {draft.duplicate_summary}",
                    severity="warning",
                ),
            )

    def _find_duplicate(
        self, draft: TransactionDraft
    ) -> tuple[Literal["deposit", "expense"], Deposit | Expense] | None:
        assert draft.transaction_date is not None and draft.amount is not None
        window_start = draft.transaction_date - timedelta(days=3)
        window_end = draft.transaction_date + timedelta(days=3)

        if draft.transaction_type == "deposit":
            stmt = select(Deposit).where(
                and_(
                    Deposit.amount == draft.amount,
                    Deposit.transaction_date >= window_start,
                    Deposit.transaction_date <= window_end,
                )
            )
            if draft.reference:
                stmt = stmt.where(
                    or_(
                        Deposit.reference == draft.reference,
                        Deposit.transaction_date == draft.transaction_date,
                    )
                )
            rows = list(self.db.scalars(stmt.limit(10)))
            best = self._best_deposit_match(draft, rows)
            return ("deposit", best) if best else None

        stmt = select(Expense).where(
            and_(
                Expense.amount == draft.amount,
                Expense.transaction_date >= window_start,
                Expense.transaction_date <= window_end,
            )
        )
        rows = list(self.db.scalars(stmt.limit(10)))
        best = self._best_expense_match(draft, rows)
        return ("expense", best) if best else None

    def _best_deposit_match(
        self, draft: TransactionDraft, rows: list[Deposit]
    ) -> Deposit | None:
        if not rows:
            return None
        scored: list[tuple[int, Deposit]] = []
        desc = (draft.description or "").lower()
        for row in rows:
            score = 0
            if row.transaction_date == draft.transaction_date:
                score += 3
            if draft.reference and row.reference and draft.reference == row.reference:
                score += 5
            row_desc = (row.description or "").lower()
            if desc and row_desc and (desc in row_desc or row_desc in desc):
                score += 2
            if draft.property_id and row.property_id == draft.property_id:
                score += 1
            scored.append((score, row))
        scored.sort(key=lambda item: item[0], reverse=True)
        return scored[0][1] if scored[0][0] >= 3 else scored[0][1]

    def _best_expense_match(
        self, draft: TransactionDraft, rows: list[Expense]
    ) -> Expense | None:
        if not rows:
            return None
        scored: list[tuple[int, Expense]] = []
        needle = " ".join(
            p for p in ((draft.vendor_name or ""), (draft.description or ""), (draft.category or "")) if p
        ).lower()
        for row in rows:
            score = 0
            if row.transaction_date == draft.transaction_date:
                score += 3
            hay = " ".join(
                p
                for p in (
                    row.vendor_name or "",
                    row.description or "",
                    row.category or "",
                )
                if p
            ).lower()
            if needle and hay and (needle in hay or hay in needle):
                score += 3
            if draft.reference and row.reference and draft.reference == row.reference:
                score += 4
            if draft.property_id and row.property_id == draft.property_id:
                score += 1
            scored.append((score, row))
        scored.sort(key=lambda item: item[0], reverse=True)
        # Prefer any amount+date window match; still flag as duplicate for user choice
        return scored[0][1]

    def _enrich_with_ai(self, drafts: list[TransactionDraft], *, buffer: Property) -> None:
        if not self.settings.llm_api_key:
            return
        catalog = []
        props = self.db.scalars(
            select(Property).options(joinedload(Property.owner))
        ).unique().all()
        for prop in props:
            catalog.append(
                {
                    "property_id": str(prop.id),
                    "client_prop_id": prop.client_prop_id,
                    "property_name": prop.name,
                    "address": prop.address,
                    "city": prop.city,
                    "owner_name": prop.owner.name if prop.owner else None,
                }
            )

        batch_size = 20
        for start in range(0, len(drafts), batch_size):
            batch = drafts[start : start + batch_size]
            try:
                self._ai_enrich_batch(batch, catalog=catalog, buffer=buffer)
            except Exception as exc:
                logger.warning(
                    "Statement AI enrichment skipped for batch starting %s: %s",
                    start,
                    exc,
                )

    def _ai_enrich_batch(
        self,
        drafts: list[TransactionDraft],
        *,
        catalog: list[dict],
        buffer: Property,
    ) -> None:
        payload_rows = []
        for draft in drafts:
            payload_rows.append(
                {
                    "row_number": draft.row_number,
                    "transaction_type": draft.transaction_type,
                    "transaction_date": (
                        draft.transaction_date.isoformat() if draft.transaction_date else None
                    ),
                    "amount": str(draft.amount) if draft.amount is not None else None,
                    "description": draft.description,
                    "reference": draft.reference,
                    "vendor_name": draft.vendor_name,
                    "current_property_id": str(draft.property_id) if draft.property_id else None,
                    "current_category": draft.category,
                    "source": draft.source,
                }
            )

        system_prompt = (
            "You enrich bank/credit-card statement rows for a property management app. "
            "Return JSON: {\"rows\":[{\"row_number\":n,\"property_id\":uuid|null,"
            "\"category\":string|null,\"is_rental_income\":bool,"
            "\"confidence\":\"high|medium|low\",\"reason\":string}]}. "
            "Pick property_id only from the catalog. "
            "If unsure, set property_id to the BUFFER property. "
            "For expenses, prefer short English categories like maintenance, utilities, "
            "tax, insurance, management_fee, other, or a short merchant label. "
            "Do not invent amounts or dates."
        )
        user_prompt = (
            f"Property catalog:\n{json.dumps(catalog, ensure_ascii=False)}\n\n"
            f"BUFFER property_id: {buffer.id}\n\n"
            f"Rows:\n{json.dumps(payload_rows, ensure_ascii=False)}"
        )

        headers = {
            "Authorization": f"Bearer {self.settings.llm_api_key}",
            "Content-Type": "application/json",
        }
        body = {
            "model": self.settings.llm_model,
            "temperature": 0,
            "response_format": {"type": "json_object"},
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
        }
        url = f"{self.settings.llm_base_url.rstrip('/')}/chat/completions"
        with httpx.Client(timeout=90.0) as client:
            response = client.post(url, headers=headers, json=body)
            response.raise_for_status()
            data = response.json()
        content = data["choices"][0]["message"]["content"]
        parsed = json.loads(content)
        rows = parsed.get("rows") if isinstance(parsed, dict) else None
        if not isinstance(rows, list):
            return

        by_row = {draft.row_number: draft for draft in drafts}
        prop_ids = {str(item["property_id"]): item for item in catalog}
        for item in rows:
            if not isinstance(item, dict):
                continue
            row_number = item.get("row_number")
            draft = by_row.get(row_number)
            if not draft:
                continue
            prop_id = item.get("property_id")
            if prop_id and str(prop_id) in prop_ids:
                meta = prop_ids[str(prop_id)]
                draft.property_id = UUID(str(prop_id))
                draft.client_prop_id = meta["client_prop_id"]
                draft.property_name = meta["property_name"]
                draft.owner_name = meta.get("owner_name")
                prop_row = self.db.get(Property, draft.property_id)
                if prop_row:
                    draft.owner_id = prop_row.owner_id
                on_buffer = meta["client_prop_id"] == BUFFER_PROP_ID
                if not on_buffer:
                    draft.warnings = [
                        w for w in draft.warnings if w.field != "property_id"
                    ]
            if draft.transaction_type == "expense" and item.get("category"):
                draft.category = str(item["category"])[:255]
            confidence = item.get("confidence")
            if confidence in {"high", "medium", "low"}:
                draft.match_confidence = confidence
            draft.status = "needs_review"
